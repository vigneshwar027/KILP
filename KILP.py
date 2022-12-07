from cmath import nan
import re
from telnetlib import ENCRYPT
import pyodbc
import os
from datetime import datetime, date
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import chardet
from xlsxwriter import Workbook
import glob

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, colors
from openpyxl.styles.colors import Color, ColorDescriptor
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from cryptography.fernet import Fernet
import warnings

warnings.filterwarnings(action='ignore')

cwd = os.path.dirname(os.path.realpath(__file__))
os.chdir(cwd)


#CHOICES
SECURE_DB_TABLES = 'OFF' #ON/OFF in caps
DB_ENCRYPTION = 'NO' #YES/NO in caps

fernet_key = b'zJD8OVkFNpd5N4fJw6pqaWiDrvybkselSQ0fF9SwXfw='
fernet = Fernet(fernet_key)

conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=localhost\SQLEXPRESS;'
                      'Database=ReportsAutomation_K;'
                      'Trusted_Connection=yes;')

cursor = conn.cursor()

def truncate_full_db():
    cursor.execute('''
                EXEC sp_MSforeachtable 'TRUNCATE TABLE ?' 
                ''')
    cursor.commit()
    # quit()


def change_format(dat):
    #print('date', date)
    
    date = str(dat).strip()

    if date in ['0100-01-01','0100/01/01','None']: 
        date = ''

    if date is nan or date =='nan':
        date = ''

    if date:
        try: 
            return datetime.strptime(date, '%m-%d-%Y').strftime('%Y-%m-%d')
        except:
            pass 
        try:
            return datetime.strptime(date, '%m-%d-%y').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%m/%d/%Y').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%m/%d/%y').strftime('%Y-%m-%d')
        except:
            pass
        ##
        try: 
            return datetime.strptime(date, '%d-%b-%Y').strftime('%Y-%m-%d')
        except:
            pass 
        try:
            return datetime.strptime(date, '%d-%b-%y').strftime('%Y-%m-%d')
        except:
            pass

        ##
        ##
        try: 
            return datetime.strptime(date, '%d-%m-%Y').strftime('%Y-%m-%d')
        except:
            pass 
        try:
            return datetime.strptime(date, '%d-%m-%y').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%d/%m/%y').strftime('%Y-%m-%d')
        except:
            pass
        ##
        try:
            return datetime.strptime(date, '%d/%b/%Y').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%d/%b/%y').strftime('%Y-%m-%d')
        except:
            pass
        
        ##
        try:
            return datetime.strptime(date, '%Y/%m/%d').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%Y-%m-%d').strftime('%Y-%m-%d')
        except:
            pass
        ##
        try:
            return datetime.strptime(date, '%Y/%b/%d').strftime('%Y-%m-%d')
        except:
            pass
        try:
            return datetime.strptime(date, '%Y-%b-%d').strftime('%Y-%m-%d')
        except:
            return ''

    else:
        return ''


def change_display_format(date):
    date = str(date).strip()
    if date:
        try:
            #'%d-%b-%y'
            return datetime.strptime(date, '%Y-%m-%d').strftime('%m/%d/%Y')
        except:
            return date

def secure_tables():
    # print('Securing data tables..')
    # # tables = ['dbo.Beneficiary','dbo.BeneficiaryPriorityDate','dbo.BeneficiaryPriorityDate','dbo.BeneficiaryEmployment','dbo.[case]']
    # tables = ['dbo.[Beneficiary]','[dbo].[Case]']
    # ben_col=['FirstName','LastName']
    # proc_col=['AccountManagerLastName','AccountManagerFirstName']
    
    # for  table in tables:
    #     table_columns = cursor.execute(''' 
    #     SELECT name FROM sys.columns WHERE object_id = OBJECT_ID('{}')
    #     '''.format(table)).fetchall()

    #     for i in range(len(table_columns)):   
    #         if (table  == '[dbo].[Case]' and (table_columns[i].name in proc_col)) or (table  == 'dbo.[Beneficiary]' and (table_columns[i].name in ben_col)): 
    #             print('hi')
    #             cursor.execute('''  UPDATE {}
    #             SET {} = NULL '''.format(table,table_columns[i].name))
    #             cursor.commit()

    ben_records = cursor.execute(''' select FirstName,BeneficiaryXref from [dbo].[beneficiary]''').fetchall()
    (len(ben_records))

    i=0
    for item  in (ben_records) :
        i+=1
        # print(item.BeneficiaryXref)
        cursor.execute(''' update [dbo].[Beneficiary]
        set FirstName = 'FN {}',
        LastName = 'LN {}' where BeneficiaryXref = '{}' '''.format(i,i,item.BeneficiaryXref))
        cursor.commit()

def process_beneficiary_file(file_path, from_name):
    # with open(file_path,'rb') as f:
    #     rawdata = b''.join([f.readline() for _ in range(20)])
    # enc= chardet.detect(rawdata)['encoding'] #UTF-16
    #print(rawdata)
    # df = pd.read_csv(file_path, encoding='utf-8',delimiter=',')
    
    df = pd.read_csv(file_path, encoding = 'unicode_escape', engine ='python',delimiter=',')

    df = df.replace("'","''", regex=True)
    df = df.replace("\.0$","", regex=True)
    # df = df.astype(str)
    list_h = df.columns.tolist()
    #print(df.iterrows())
    total_rows = len(df)
    for index, row in df.iterrows():
        #print(index)
        #print(row['organization_group_id'])
        #if(index==3):
            #break
            #return False
        if True:
            organization_xref = ''
            if 'organization_group_id' in list_h and not pd.isna(row['organization_group_id']):
                organization_xref = str(row['organization_group_id']).strip().replace('.0','')
                # organization_xref = fernet.encrypt(organization_xref.encode())
        

            organization_name = ''
            if 'organization_group_name' in list_h and not pd.isna(row['organization_group_name']):
                organization_name = str(str(row['organization_group_name'])).strip()
                if DB_ENCRYPTION == 'YES':
                    organization_name = (fernet.encrypt(organization_name.encode())).decode('utf-8')

            if organization_xref and organization_name:
                results = cursor.execute(
                    '''select distinct * FROM dbo.Organization where OrganizationXref='{}' '''.format(organization_xref)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute(
                        ''' INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}') '''.format(
                            organization_xref, organization_name))
                    cursor.commit()


            petitioner_xref = ''
            if 'petitioner_company_id' in list_h and not pd.isna(row['petitioner_company_id']):
                petitioner_xref = str(row['petitioner_company_id']).strip().replace('.0','')
            # print(petitioner_xref)
            petitioner_name = ''
            if 'petitioner_company_name' in list_h and not pd.isna(row['petitioner_company_name']):
                petitioner_name = str(str(row['petitioner_company_name'])).strip()

            petitioner_company_of_primary_beneficiary = ''
            if 'petitioner_company_of_primary_beneficiary' in list_h and not pd.isna(row['petitioner_company_of_primary_beneficiary']):
                petitioner_company_of_primary_beneficiary = str(
                    str(row['petitioner_company_of_primary_beneficiary'])).strip()
        #     print(petitioner_company_of_primary_beneficiary)
        # quit()
        # if True:
            beneficiary_xref = ''
            if 'beneficiary_id' in list_h and not pd.isna(row['beneficiary_id']):
                beneficiary_xref = str(row['beneficiary_id']).strip().replace('.0','')

            beneficiary_xref2 = ''
            if 'fn_case_id' in list_h and not pd.isna(row['fn_case_id']):
                beneficiary_xref2 = str(row['fn_case_id']).strip().replace('.0','')

            beneficiary_type = ''
            if 'beneficiary_type' in list_h and not pd.isna(row['beneficiary_type']):
                beneficiary_type = str(row['beneficiary_type']).strip()
                
            is_primary_beneficiary = 1

            if petitioner_xref and petitioner_name:
                petitioners = cursor.execute('''select distinct * from dbo.Petitioner where PetitionerXref='{}' '''.format(petitioner_xref)).fetchall()
                if len(petitioners) <= 0 :
                    cursor.execute( 
                            '''  INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName) VALUES ('{}', '{}') '''.format(
                                petitioner_xref, petitioner_name))
                    cursor.commit()


                # if beneficiary_type == 'Dependent':
                #     if str(row['primary_beneficiary_id']).strip():
                #         ##print('SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}''.format(row['Primary Beneficiary Xref'].strip()))
                #         results = cursor.execute(
                #             'SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}''.format(
                #                 str(row['primary_beneficiary_id']).strip())).fetchall()
                #         length = len(results)
                #         if length > 0:
                #             petitioner_id = results[0][0]
                #             print(results[0])
                #             # quit()
                #             is_primary_beneficiary = 0
                        

                # else:
                #     ##print('SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and PetitionerName = '{}' and OrganizationId={}'.format(petitioner_xref, petitioner_name, organization_id))
                #     results = cursor.execute(
                #         'SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and OrganizationId={}'.format(
                #             petitioner_xref, organization_id)).fetchall()
                #     length = len(results)
                #     if length <= 0:
                #         ##print('INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')'.format(petitioner_xref, petitioner_name, organization_id))
                #         cursor.execute(
                #             'INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')'.format(
                #                 petitioner_xref, petitioner_name, organization_id))
                #         cursor.execute('SELECT @@IDENTITY AS ID;')
                #         petitioner_id = cursor.fetchone()[0]
                #         cursor.commit()
                #     else:
                #         petitioner_id = results[0].PetitionerId

            if True: # if petitioner id condition is replaced as if true
                

                beneficiary_record_creation_date = ''
                if 'beneficiary_record_opened_date' in list_h and str(
                        row['beneficiary_record_opened_date']).strip() and not pd.isna(row['beneficiary_record_opened_date']):
                    beneficiary_record_creation_date = change_format(row['beneficiary_record_opened_date'])

                beneficiary_record_inactivation_date = None
                if 'beneficiary_record_retired_date' in list_h and str(
                        row['beneficiary_record_retired_date']).strip() and not pd.isna(
                        row['beneficiary_record_retired_date']):
                    beneficiary_record_inactivation_date = change_format(row['beneficiary_record_retired_date'])

                beneficiary_record_status = 0
                if 'beneficiary_status' in list_h and not pd.isna(row['beneficiary_status']):
                    beneficiary_record_status_chk = str(row['beneficiary_status']).strip()
                    if beneficiary_record_status_chk == 'Active':
                        beneficiary_record_status = 1

                beneficiary_employee_id = ''
                if 'employee_id' in list_h and not pd.isna(row['employee_id']):
                    beneficiary_employee_id = str(row['employee_id']).strip()

                beneficiary_last_name = ''
                if 'beneficiary_last_name' in list_h and not pd.isna(row['beneficiary_last_name']):
                    beneficiary_last_name = str(str(row['beneficiary_last_name']).strip())

                beneficiary_first_name = ''
                if 'beneficiary_first_name' in list_h and not pd.isna(row['beneficiary_first_name']):
                    beneficiary_first_name = str(str(row['beneficiary_first_name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode('utf-8')

                beneficiary_middle_name = ''
                if 'Beneficiary Middle Name' in list_h and not pd.isna(row['Beneficiary Middle Name']):
                    beneficiary_middle_name = str(str(row['Beneficiary Middle Name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        beneficiary_middle_name = (fernet.encrypt(beneficiary_middle_name.encode())).decode('utf-8')

                primary_beneficiary_id = ''
                if 'primary_beneficiary_id' in list_h and not pd.isna(row['primary_beneficiary_id']):
                    primary_beneficiary_id = str(row['primary_beneficiary_id']).strip().replace('.0','')

                # print(primary_beneficiary_id)
                if primary_beneficiary_id == beneficiary_xref:
                    is_primary_beneficiary = 1
                else:
                    is_primary_beneficiary = 0

                primary_beneficiary_last_name = ''
                if 'primary_beneficiary_last_name' in list_h and not pd.isna(row['primary_beneficiary_last_name']):
                    primary_beneficiary_last_name = str(str(row['primary_beneficiary_last_name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode(
                            'utf-8')

                primary_beneficiary_first_name = ''
                if 'primary_beneficiary_first_name' in list_h and not pd.isna(row['primary_beneficiary_first_name']):
                    primary_beneficiary_first_name = str(str(row['primary_beneficiary_first_name']).strip())

                    if DB_ENCRYPTION == 'YES':
                        primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode(
                            'utf-8')

                relation = ''
                if 'relation' in list_h and not pd.isna(row['relation']):
                    relation = str(str(row['relation']).strip())
                    if DB_ENCRYPTION == 'YES':
                        relation = (fernet.encrypt(relation.encode())).decode('utf-8')

                gender = ''
                if 'gender' in list_h and not pd.isna(row['gender']):
                    gender = str(str(row['gender']).strip())
                    if DB_ENCRYPTION == 'YES':
                        gender = (fernet.encrypt(gender.encode())).decode('utf-8')

                date_of_birth = ''
                if 'date_of_birth' in list_h and row['date_of_birth'] and not pd.isna(row['date_of_birth']):
                    date_of_birth = change_format(row['date_of_birth'])

                country_of_birth = ''
                if 'country_of_birth' in list_h and not pd.isna(row['country_of_birth']):
                    country_of_birth = str(str(row['country_of_birth']).strip())
                    if DB_ENCRYPTION == 'YES':
                        country_of_birth = (fernet.encrypt(country_of_birth.encode())).decode('utf-8')

                country_of_citizenship = ''
                if 'country_of_citizenship' in list_h and not pd.isna(row['country_of_citizenship']):
                    country_of_citizenship = row['country_of_citizenship']
                    if DB_ENCRYPTION == 'YES':
                        country_of_citizenship = (fernet.encrypt(country_of_citizenship.encode())).decode('utf-8')

                alien_number = ''
                if 'alien_number' in list_h and not pd.isna(row['alien_number']):
                    alien_number = str(row['alien_number']).strip().replace('.0','')
                    if DB_ENCRYPTION == 'YES':
                        alien_number = (fernet.encrypt(str(alien_number).encode())).decode('utf-8')

                date_of_last_entry_into_the_us = ''
                if 'date_of_last_entry_into_the_us' in list_h and str(
                        row['date_of_last_entry_into_the_us']).strip() and not pd.isna(
                        row['date_of_last_entry_into_the_us']):
                    date_of_last_entry_into_the_us = change_format(row['date_of_last_entry_into_the_us'])

                i94_number = ''
                if 'i94_number' in list_h and not pd.isna(row['i94_number']):
                    i94_number = str(row['i94_number']).strip()
                    if DB_ENCRYPTION == 'YES':
                        i94_number = (fernet.encrypt(i94_number.encode())).decode('utf-8')

                immigration_status = ''
                if 'i94_status' in list_h and not pd.isna(row['i94_status']):
                    immigration_status = row['i94_status']

                immigration_status_valid_from = ''
                if 'i94_valid_from' in list_h and str(row['i94_valid_from']).strip() and not pd.isna(row['i94_valid_from']):
                    
                    immigration_status_valid_from = change_format(row['i94_valid_from'])
                    

                immigration_status_expiration_status = ''
                if 'i94_exp_date' in list_h and str(row['i94_exp_date']).strip() and not pd.isna(row['i94_exp_date']):
                    if str(row['i94_exp_date']).strip() == 'D/S':
                        immigration_status_expiration_status = 'D/S'
                    else:
                        if 'D/S' in str(row['i94_exp_date']).strip():
                            split1 = (str(row['i94_exp_date']).strip()).split('(D/S)')
                            ##print(split1)
                            immigration_status_expiration_status = change_format(split1[0])
                            immigration_status_expiration_status = str(immigration_status_expiration_status) + ' (D/S)'
                        else:
                            immigration_status_expiration_status = change_format(str(row['i94_exp_date']).strip())

                # print(immigration_status_expiration_status)
                i797_approved_date = ''
                if 'I-797 Approved Date' in list_h and str(row['I-797 Approved Date']).strip() and not pd.isna(
                        row['I-797 Approved Date']):
                    i797_approved_date = change_format(row['I-797 Approved Date'])

                i797_status = ''
                if 'I-797 Status' in list_h and not pd.isna(row['I-797 Status']):
                    i797_status = str(row['I-797 Status']).strip()
                    if DB_ENCRYPTION == 'YES':
                        i797_status = (fernet.encrypt(i797_status.encode())).decode('utf-8')

                i797_valid_from = ''
                if 'i797_valid_from' in list_h and str(row['i797_valid_from']).strip() and not pd.isna(
                        row['i797_valid_from']):
                    i797_valid_from = change_format(str(row['i797_valid_from']))

                i797_expiration_date = ''
                if 'i797_exp_date' in list_h and str(row['i797_exp_date']).strip() and not pd.isna(row['i797_exp_date']):
                    i797_expiration_date = change_format(str(row['i797_exp_date']))

                final_niv_status_valid_from = ''
                if '#' in list_h and str(row['final_niv_hl_status_valid_from']).strip() and not pd.isna(
                        row['final_niv_hl_status_valid_from']):
                    final_niv_status_valid_from = change_format(row['final_niv_hl_status_valid_from'])

                final_niv_maxout_date = ''
                if 'final_niv_maxout_date' in list_h and str(row['final_niv_maxout_date']).strip() and not pd.isna(
                        row['final_niv_maxout_date']):
                    final_niv_maxout_date = change_format(row['final_niv_maxout_date'])

                #print('final_niv_maxout_date ', final_niv_maxout_date)
                maxout_note = ''
                if 'Maxout Date Applicability and Note' in list_h and not pd.isna(
                        row['Maxout Date Applicability and Note']):
                    maxout_note = str(str(row['Maxout Date Applicability and Note']).strip())
                    if DB_ENCRYPTION == 'YES':
                        maxout_note = (fernet.encrypt(maxout_note.encode())).decode('utf-8')

                ped = ''
                if 'ped_petition_end_date' in list_h and str(row['ped_petition_end_date']).strip() and not pd.isna(
                        row['ped_petition_end_date']):
                    ped = change_format(row['ped_petition_end_date'])

                ead_type = ''
                if 'ead_type' in list_h and not pd.isna(row['ead_type']):
                    ead_type = str(str(row['ead_type']).strip())
                    if DB_ENCRYPTION == 'YES':
                        ead_type = (fernet.encrypt(ead_type.encode())).decode('utf-8')

                ead_valid_from = ''
                if 'EAD Valid From' in list_h and str(row['EAD Valid From']).strip() and not pd.isna(row['EAD Valid From']):
                    ead_valid_from = change_format(row['EAD Valid From'])

                ead_expiration_date = ''
                if 'ead_exp_date' in list_h and str(row['ead_exp_date']).strip() and not pd.isna(row['ead_exp_date']):
                    ead_expiration_date = change_format(row['ead_exp_date'])
                
                ap_valid_from = ''
                if 'AP Valid From' in list_h and row['AP Valid From'].strip() and not pd.isna(row['AP Valid From']):
                    ap_valid_from = change_format(row['AP Valid From'])

                ap_expiration_date = ''
                if 'ap_exp_date' in list_h and str(row['ap_exp_date']).strip() and not pd.isna(row['ap_exp_date']):
                    ap_expiration_date = change_format(row['ap_exp_date'])

                ead_ap_type = ''
                if 'EAD/AP Type' in list_h and not pd.isna(row['EAD/AP Type']):
                    ead_ap_type = str(row['EAD/AP Type']).strip()
                    if DB_ENCRYPTION == 'YES':
                        ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode('utf-8')

                ead_ap_valid_from = ''
                if 'EAD/AP Valid From' in list_h and str(row['EAD/AP Valid From']).strip() and not pd.isna(
                        row['EAD/AP Valid From']):
                    ead_ap_valid_from = change_format(row['EAD/AP Valid From'])

                ead_ap_expiration_date = ''
                if 'EAD/AP Expiration Date' in list_h and str(row['EAD/AP Expiration Date']).strip() and not pd.isna(
                        row['EAD/AP Expiration Date']):
                    ead_ap_expiration_date = change_format(row['EAD/AP Expiration Date'])

                ds_2019_valid_from = ''
                if 'ds2019_valid_from' in list_h and str(row['ds2019_valid_from']).strip() and not pd.isna(
                        row['ds2019_valid_from']):
                    ds_2019_valid_from = change_format(row['ds2019_valid_from'])

                ds_2019_expiration_date = ''
                if 'ds2019_exp_date' in list_h and str(row['ds2019_exp_date']).strip() and not pd.isna(
                        row['ds2019_exp_date']):
                    ds_2019_expiration_date = change_format(row['ds2019_exp_date'])

                reentry_permit_expiration_date = ''
                if 're_entry_permit_exp_date' in list_h and row['re_entry_permit_exp_date'] and not pd.isna(
                        row['re_entry_permit_exp_date']):
                    reentry_permit_expiration_date = change_format(row['re_entry_permit_exp_date'])

                green_card_valid_from = ''
                if 'Green Card Valid From' in list_h and row['Green Card Valid From'] and not pd.isna(
                        row['Green Card Valid From']):
                    green_card_valid_from = change_format(row['Green Card Valid From'])

                green_card_expiration_date = ''
                if 'green_card_exp_date' in list_h and row['green_card_exp_date'] and not pd.isna(
                        row['green_card_exp_date']):
                    green_card_expiration_date = change_format(row['green_card_exp_date'])

                passport_last_name = ''
                if 'Passport Last Name' in list_h and not pd.isna(row['Passport Last Name']):
                    passport_last_name = str(str(row['Passport Last Name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        passport_last_name = (fernet.encrypt(passport_last_name.encode())).decode('utf-8')

                passport_first_name = ''
                if 'Passport First Name' in list_h and not pd.isna(row['Passport First Name']):
                    passport_first_name = str(str(row['Passport First Name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        passport_first_name = (fernet.encrypt(passport_first_name.encode())).decode('utf-8')

                passport_middle_name = ''
                if 'Passport Middle Name' in list_h and not pd.isna(row['Passport Middle Name']):
                    passport_middle_name = str(str(row['Passport Middle Name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        passport_middle_name = (fernet.encrypt(passport_middle_name.encode())).decode('utf-8')

                passport_number = ''
                if 'passport_number' in list_h and not pd.isna(row['passport_number']):
                    passport_number = str(row['passport_number']).strip()
                    if DB_ENCRYPTION == 'YES':
                        passport_number = (fernet.encrypt(passport_number.encode())).decode('utf-8')

                passport_issuing_country = ''
                if 'passport_issuing_country' in list_h and not pd.isna(row['passport_issuing_country']):
                    passport_issuing_country = str(str(row['passport_issuing_country']).strip())
                    if DB_ENCRYPTION == 'YES':
                        passport_issuing_country = (fernet.encrypt(passport_issuing_country.encode())).decode('utf-8')

                passport_valid_from = ''
                if 'passport_valid_from' in list_h and row['passport_valid_from'] and not pd.isna(row['passport_valid_from']):
                    passport_valid_from = change_format(row['passport_valid_from'])
                    if DB_ENCRYPTION == 'YES':
                        ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode('utf-8')

                passport_expiration_date = ''
                if 'passport_exp_date' in list_h and row['passport_exp_date'] and not pd.isna(row['passport_exp_date']):
                    passport_expiration_date = change_format(row['passport_exp_date'])

                visa_type = ''
                if 'visa_type' in list_h and not pd.isna(row['visa_type']):
                    visa_type = str(row['visa_type']).strip()
                    if DB_ENCRYPTION == 'YES':
                        visa_type = (fernet.encrypt(visa_type.encode())).decode('utf-8')

                visa_valid_from = ''
                if 'visa_issue_date' in list_h and row['visa_issue_date'] and not pd.isna(row['visa_issue_date']):
                    visa_valid_from = change_format(row['visa_issue_date'])

                visa_expiration_date = ''
                if 'visa_exp_date' in list_h and row['visa_exp_date'] and not pd.isna(row['visa_exp_date']):
                    visa_expiration_date = change_format(row['visa_exp_date'])

                employee_hire_date = ''
                if 'hire_date' in list_h and row['hire_date'] and not pd.isna(row['hire_date']):
                    employee_hire_date = change_format(row['hire_date'])

                current_job_title = ''
                if 'job_title' in list_h and not pd.isna(row['job_title']):
                    current_job_title = str(str(row['job_title']).strip())
                    if DB_ENCRYPTION == 'YES':
                        current_job_title = (fernet.encrypt(current_job_title.encode())).decode('utf-8')

                work_address_street = ''
                if 'job_location_street' in list_h and not pd.isna(row['job_location_street']):
                    work_address_street = str(str(row['job_location_street']).strip())
                    if DB_ENCRYPTION == 'YES':
                        work_address_street = (fernet.encrypt(work_address_street.encode())).decode('utf-8')

                work_address_city = ''
                if 'job_location_city' in list_h and not pd.isna(row['job_location_city']):
                    work_address_city = str(str(row['job_location_city']).strip())
                    if DB_ENCRYPTION == 'YES':
                        work_address_city = (fernet.encrypt(work_address_city.encode())).decode('utf-8')

                work_address_state = ''
                if 'job_location_state' in list_h and not pd.isna(row['job_location_state']):
                    work_address_state = str(str(row['job_location_state']).strip())
                    if DB_ENCRYPTION == 'YES':
                        work_address_state = (fernet.encrypt(work_address_state.encode())).decode('utf-8')

                work_address_zip = ''
                if 'Work Address-Zip' in list_h and not pd.isna(row['Work Address-Zip']):
                    work_address_zip = str(str(row['Work Address-Zip']).strip())
                    if DB_ENCRYPTION == 'YES':
                        work_address_zip = (fernet.encrypt(work_address_zip.encode())).decode('utf-8')

                work_address_country = ''
                if 'Work Address-Country' in list_h and not pd.isna(row['Work Address-Country']):
                    work_address_country = str(row['Work Address-Country'].strip())
                    if DB_ENCRYPTION == 'YES':
                        work_address_country = (fernet.encrypt(work_address_country.encode())).decode('utf-8')

                priority_date_1_date = ''
                if 'primary_visa_priority_date' in list_h and row['primary_visa_priority_date'] and not pd.isna(
                        row['primary_visa_priority_date']):
                    priority_date_1_date = change_format(row['primary_visa_priority_date'])

                priority_date_1_category = ''
                if 'primary_visa_priority_category' in list_h and not pd.isna(row['primary_visa_priority_category']):
                    priority_date_1_category = str(str(row['primary_visa_priority_category']).strip())

                priority_date_1_country_of_charge = ''
                if 'primary_visa_country_of_chargeability' in list_h and not pd.isna(
                        row['primary_visa_country_of_chargeability']):
                    priority_date_1_country_of_charge = str(row['primary_visa_country_of_chargeability']).strip()

                priority_date_2_date = ''
                if 'Priority Date 2-Date' in list_h and row['Priority Date 2-Date'].strip() and not pd.isna(
                        row['Priority Date 2-Date']):
                    priority_date_2_date = change_format(row['Priority Date 2-Date'])

                priority_date_2_category = ''
                if 'Priority Date 2-Category' in list_h and not pd.isna(row['Priority Date 2-Category']):
                    priority_date_2_category = str(str(row['Priority Date 2-Category']).strip())

                priority_date_2_country_of_charge = ''
                if 'Priority Date 2-Country of Chargeability' in list_h and not pd.isna(
                        row['Priority Date 2-Country of Chargeability']):
                    priority_date_2_country_of_charge = str(
                        str(row['Priority Date 2-Country of Chargeability']).strip())

                priority_date_3_date = ''
                if 'Priority Date 3-Date' in list_h and row['Priority Date 3-Date'].strip() and not pd.isna(
                        row['Priority Date 3-Date']):
                    priority_date_3_date = change_format(row['Priority Date 3-Date'])

                priority_date_3_category = ''
                if 'Priority Date 3-Category' in list_h and not pd.isna(row['Priority Date 3-Category']):
                    priority_date_3_category = str(str(row['Priority Date 3-Category']).strip())

                priority_date_3_country_of_charge = ''
                if 'Priority Date 3-Country of Chargeability' in list_h and not pd.isna(
                        row['Priority Date 3-Country of Chargeability']):
                    priority_date_3_country_of_charge = str(
                        str(row['Priority Date 3-Country of Chargeability']).strip())

                priority_date_4_date = ''
                if 'Priority Date 4-Date' in list_h and row['Priority Date 4-Date'] and not pd.isna(
                        row['Priority Date 4-Date']):
                    priority_date_4_date = change_format(row['Priority Date 4-Date'])

                priority_date_4_category = ''
                if 'Priority Date 4-Category' in list_h and not pd.isna(row['Priority Date 4-Category']):
                    priority_date_4_category = str(str(row['Priority Date 4-Category']).strip())

                priority_date_4_country_of_charge = ''
                if 'Priority Date 4-Country of Chargeability' in list_h and not pd.isna(
                        row['Priority Date 4-Country of Chargeability']):
                    priority_date_4_country_of_charge = str(
                        str(row['Priority Date 4-Country of Chargeability']).strip())

                priority_date_5_date = ''
                if 'Priority Date 5-Date' in list_h and row['Priority Date 5-Date'] and not pd.isna(
                        row['Priority Date 5-Date']):
                    priority_date_5_date = change_format(row['Priority Date 5-Date'])

                priority_date_5_category = ''
                if 'Priority Date 5-Category' in list_h and not pd.isna(row['Priority Date 5-Category']):
                    priority_date_5_category = str(str(row['Priority Date 5-Category']).strip())

                priority_date_5_country_of_charge = ''
                if 'Priority Date 5-Country of Chargeability' in list_h and not pd.isna(
                        row['Priority Date 5-Country of Chargeability']):
                    priority_date_5_country_of_charge = str(
                        str(row['Priority Date 5-Country of Chargeability']).strip())

               

                FullName = ''
                if 'beneficiary_full_name' in list_h and not pd.isna(
                        row['beneficiary_full_name']):
                    FullName = str(str(row['beneficiary_full_name'])).strip()

                ImmigrationStatusExpirationDate2 = ''
                if 'current_status_exp_date' in list_h and str(row['current_status_exp_date']).strip() and not pd.isna(
                        row['current_status_exp_date']):
                    ImmigrationStatusExpirationDate2 = change_format(row['current_status_exp_date'])

                I129SEndDate = ''
                if 'i129s_exp_date' in list_h and str(row['i129s_exp_date']).strip() and not pd.isna(row['i129s_exp_date']):
                    I129SEndDate = change_format(row['i129s_exp_date'])

                GreenCardMethod = ''
                if 'green_card_method' in list_h and not pd.isna(row['green_card_method']):
                    GreenCardMethod = str(row['green_card_method']).strip()

                WorkEmail = ''
                if 'work_email_id' in list_h and not pd.isna(row['work_email_id']):
                    WorkEmail = str(row['work_email_id']).strip()

                current_employer = ''
                if 'current_employer' in list_h and not pd.isna(row['current_employer']):
                    current_employer = str(row['current_employer']).strip()

                EmployeeId = ''
                if 'management_info_employee_id' in list_h and not pd.isna(row['management_info_employee_id']):
                    EmployeeId = str(row['management_info_employee_id']).strip()

                Department = ''
                if 'management_info_department' in list_h and not pd.isna(row['management_info_department']):
                    Department = str(row['management_info_department']).strip()

                Department_Group = ''
                if 'management_info_dept_group' in list_h and not pd.isna(row['management_info_dept_group']):
                    Department_Group = str(row['management_info_dept_group']).strip()

                Department_Number = ''
                if 'management_info_dept_number' in list_h and not pd.isna(row['management_info_dept_number']):
                    Department_Number = str(row['management_info_dept_number']).strip()

                Business_Unit_Code = ''
                if 'business_unit_code' in list_h and not pd.isna(row['business_unit_code']):
                    Business_Unit_Code = str(row['business_unit_code']).strip()

                Client_Billing_Code = ''
                if 'hr_info_client_billing_code' in list_h and not pd.isna(row['hr_info_client_billing_code']):
                    Client_Billing_Code = str(row['hr_info_client_billing_code']).strip()
                    Client_Billing_Code = Client_Billing_Code

                ManagerName = ''
                if 'management_info_manager' in list_h and not pd.isna(row['management_info_manager']):
                    ManagerName = str(str(row['management_info_manager'])).strip()

                ManagerEmail = ''
                if 'management_info_manager_email' in list_h and not pd.isna(row['management_info_manager_email']):
                    ManagerEmail = str(row['management_info_manager_email']).strip()

                SecondLevelManager = ''
                if 'management_info_second_level_manager' in list_h and not pd.isna(row['management_info_second_level_manager']):
                    SecondLevelManager = str(str(row['management_info_second_level_manager'])).strip()

                SecondLevelManagerEmail = ''
                if 'management_info_second_level_manager_email' in list_h and not pd.isna(row['management_info_second_level_manager_email']):
                    SecondLevelManagerEmail = str(row['management_info_second_level_manager_email']).strip()

                BusinessPartnerName = ''
                if 'management_info_partner_name' in list_h and not pd.isna(row['management_info_partner_name']):
                    BusinessPartnerName = str(str(row['management_info_partner_name'])).strip()

                BusinessPartnerEmail = ''
                if 'management_info_partner_email' in list_h and not pd.isna(row['management_info_partner_email']):
                    BusinessPartnerEmail = str(row['management_info_partner_email']).strip()

                CostCenter = ''
                if 'management_info_cost_center' in list_h and not pd.isna(row['management_info_cost_center']):
                    CostCenter = str(row['management_info_cost_center']).strip()

                CostCenterNumber = ''
                if 'management_info_cost_center_number' in list_h and not pd.isna(row['management_info_cost_center_number']):
                    CostCenterNumber = str(row['management_info_cost_center_number']).strip()

                ClientBillingCode = ''
                if 'management_info_client_billing_code' in list_h and not pd.isna(row['management_info_client_billing_code']):
                    ClientBillingCode = str(row['management_info_client_billing_code']).strip()
                    ClientBillingCode = ClientBillingCode

                BusinessUnitCode = ''
                if 'management_info_business_unit_code' in list_h and not pd.isna(row['management_info_business_unit_code']):
                    BusinessUnitCode = str(row['management_info_business_unit_code']).strip()

                JobTitle = ''
                if 'management_info_job_title' in list_h and not pd.isna(row['management_info_job_title']):
                    JobTitle = str(str(row['management_info_job_title'])).strip()

                JobCode = ''
                if 'management_info_job_code' in list_h and not pd.isna(row['management_info_job_code']):
                    JobCode = str(row['management_info_job_code']).strip()

                EmploymentStartDate = ''
                if 'management_info_job_start_date' in list_h and str(
                        row['management_info_job_start_date']).strip() and not pd.isna(
                        row['management_info_job_start_date']):
                    EmploymentStartDate = change_format(row['management_info_job_start_date'])

                EmploymentEndDate = ''
                if 'management_info_job_end_date' in list_h and str(
                        row['management_info_job_end_date']).strip() and not pd.isna(
                        row['management_info_job_end_date']):
                    EmploymentEndDate = change_format(row['management_info_job_end_date'])

                WorkAddressFull = ''
                if 'management_info_work_address' in list_h and not pd.isna(row['management_info_work_address']):
                    WorkAddressFull = str(str(row['management_info_work_address'])).strip()

                WorkLocationCity = ''
                if 'management_info_job_location_city' in list_h and not pd.isna(row['management_info_job_location_city']):
                    WorkLocationCity = str(row['management_info_job_location_city']).strip()

                WorkLocationState = ''
                if 'management_info_job_location_state' in list_h and not pd.isna(row['management_info_job_location_state']):
                    WorkLocationState = str(row['management_info_job_location_state']).strip()
                
                WorkLocationCountry = ''
                if 'management_info_job_location_country' in list_h and not pd.isna(row['management_info_job_location_country']):
                    WorkLocationCountry = str(row['management_info_job_location_country']).strip()

                Visa_GreenCardMethod = ''
                if 'visa_priority_green_card_method' in list_h and not pd.isna(row['visa_priority_green_card_method']):
                    Visa_GreenCardMethod = str(row['visa_priority_green_card_method']).strip()

                PriorityDate1Note = ''
                if 'visa_priority_note' in list_h and not pd.isna(row['visa_priority_note']):
                    PriorityDate1Note = str(str(row['visa_priority_note'])).strip()

                Current_Immigration_Status = ''
                if 'current_status' in list_h:
                    Current_Immigration_Status =  str(row['current_status']).strip()

                if beneficiary_xref:
                    results = cursor.execute('''select distinct * FROM dbo.Beneficiary where BeneficiaryXref='{}' '''.format(beneficiary_xref)).fetchall()
                    length = len(results)
                    if length <= 0:

                        cursor.execute('''INSERT INTO dbo.Beneficiary(OrganizationXref,PetitionerofPrimaryBeneficiary, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, MiddleName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, Gender, BirthDate, BirthCountry, CitizenshipCountry, AlienNumber, MostRecentUSEntryDate, I94Number, ImmigrationStatus, ImmigrationStatusValidFromDate, ImmigrationStatusExpirationDate, MostRecentI797IssueApprovalDate, MostRecentI797Status, MostRecentI797ValidFromDate, I797ExpirationDate, InitialHlEntryDate, FinalNivDate, MaxOutDateNote, EadType, VisaPedDate, EadValidFromDate, EadExpirationDate, AdvanceParoleValidFromDate, AdvanceParoleExpirationDate, EADAPType, EadApValidFromDate, EadApExpirationDate, Ds2019ValidFromDate, Ds2019ExpirationDate, ReEntryPermitExpirationDate, GreenCardValidFromDate, GreenCardExpirationDate, MostRecentPassportLastName, MostRecentPassportFirstName, MostRecentPassportMiddleName, MostRecentPassportNumber, MostRecentPassportIssuingCountry, MostRecentPassportValidFromDate, MostRecentPassportExpirationDate, VisaType, VisaValidFromDate, VisaExpirationDate, from_name, is_primary_beneficiary,Beneficiary_Xref2,FullName,Current_Immigration_Status,CurrentImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note,PetitionerXref,PriorityDate1Date,PriorityDate1Category,PrioritDate1Country) VALUES ('{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}','{}','{}','{}','{}') '''.format(organization_xref,petitioner_company_of_primary_beneficiary, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary,beneficiary_xref2,FullName,Current_Immigration_Status,ImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note,petitioner_xref,priority_date_1_date,priority_date_1_category,priority_date_1_country_of_charge))
                        cursor.commit()
                        
                    else:
                        
                        cursor.execute('''  UPDATE dbo.Beneficiary SET OrganizationXref = '{}',PetitionerXref='{}',PetitionerofPrimaryBeneficiary='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', MiddleName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', Gender='{}', BirthDate='{}', BirthCountry='{}', CitizenshipCountry='{}', AlienNumber='{}', MostRecentUSEntryDate='{}', I94Number='{}', ImmigrationStatus='{}', ImmigrationStatusValidFromDate='{}', ImmigrationStatusExpirationDate='{}', MostRecentI797IssueApprovalDate='{}', MostRecentI797Status='{}', MostRecentI797ValidFromDate='{}', I797ExpirationDate='{}', InitialHlEntryDate='{}', FinalNivDate='{}', MaxOutDateNote='{}', EadType='{}', VisaPedDate='{}', EadValidFromDate='{}', EadExpirationDate='{}', AdvanceParoleValidFromDate='{}', AdvanceParoleExpirationDate='{}', EADAPType='{}', EadApValidFromDate='{}', EadApExpirationDate='{}', Ds2019ValidFromDate='{}', Ds2019ExpirationDate='{}', ReEntryPermitExpirationDate='{}', GreenCardValidFromDate='{}', GreenCardExpirationDate='{}', MostRecentPassportLastName='{}', MostRecentPassportFirstName='{}', MostRecentPassportMiddleName='{}', MostRecentPassportNumber='{}', MostRecentPassportIssuingCountry='{}', MostRecentPassportValidFromDate='{}', MostRecentPassportExpirationDate='{}', VisaType='{}', VisaValidFromDate='{}', VisaExpirationDate='{}', from_name='{}', is_primary_beneficiary='{}',Beneficiary_Xref2='{}',FullName='{}',Current_Immigration_Status='{}',CurrentImmigrationStatusExpirationDate2='{}',I129SEndDate='{}',GreenCardMethod='{}',WorkEmail='{}',current_employer='{}',Visa_GreenCardMethod='{}',PriorityDate1Note='{}',PriorityDate1Date='{}',PriorityDate1Category='{}',PrioritDate1Country='{}' WHERE BeneficiaryXref='{}' '''.format(organization_xref,petitioner_xref,petitioner_company_of_primary_beneficiary, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary,beneficiary_xref2,FullName,Current_Immigration_Status,ImmigrationStatusExpirationDate2,I129SEndDate,GreenCardMethod,WorkEmail,current_employer,Visa_GreenCardMethod,PriorityDate1Note,priority_date_1_date,priority_date_1_category,priority_date_1_country_of_charge,beneficiary_xref))
                        cursor.commit()
                
        
                if beneficiary_xref:
                    results = cursor.execute('''select distinct * FROM dbo.BeneficiaryPriorityDate where BeneficiaryXref='{}' '''.format(beneficiary_xref)).fetchall()
                    length = len(results)
                    if length <= 0:
                        cursor.execute(''' INSERT INTO dbo.BeneficiaryPriorityDate(BeneficiaryXref, Priority1Date, Priority1Category, Priority1Country, Priority2Date, Priority2Category, Priority2Country, Priority3Date, Priority3Category, Priority3Country, Priority4Date, Priority4Category, Priority4Country, Priority5Date, Priority5Category, Priority5Country) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}') '''.format(beneficiary_xref, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge))
                        cursor.commit()
                    else:
                        cursor.execute('''  UPDATE dbo.BeneficiaryPriorityDate SET BeneficiaryXref='{}', Priority1Date='{}', Priority1Category='{}', Priority1Country='{}', Priority2Date='{}', Priority2Category='{}', Priority2Country='{}', Priority3Date='{}', Priority3Category='{}', Priority3Country='{}', Priority4Date='{}', Priority4Category='{}', Priority4Country='{}', Priority5Date='{}', Priority5Category='{}', Priority5Country='{}' WHERE BeneficiaryXref='{}' '''.format(beneficiary_xref, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge,beneficiary_xref))
                        cursor.commit()

                if beneficiary_xref:
                    results = cursor.execute('''  SELECT * FROM dbo.BeneficiaryEmployment where BeneficiaryXref='{}' '''.format(beneficiary_xref)).fetchall()
                    length = len(results)
                    if length <= 0:
                        cursor.execute('''  INSERT INTO dbo.BeneficiaryEmployment(BeneficiaryXref, EmployeeId, HireDate, JobTitle, Address1, City, StateProvince, ZipCode, Country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState,WorkLocationCountry) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}') '''.format(beneficiary_xref, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState,WorkLocationCountry))
                        cursor.commit()
                    else:
                        cursor.execute('''  UPDATE dbo.BeneficiaryEmployment SET WorkLocationCountry = '{}',BeneficiaryXref='{}', EmployeeId='{}', HireDate='{}', JobTitle='{}', Address1='{}', City='{}', StateProvince='{}', ZipCode='{}', Country='{}',Department='{}',Department_Group='{}',Department_Number='{}',Business_Unit_Code='{}',Client_Billing_Code='{}',ManagerName='{}',ManagerEmail='{}',SecondLevelManager='{}',SecondLevelManagerEmail='{}',BusinessPartnerName='{}',BusinessPartnerEmail='{}',CostCenter='{}',CostCenterNumber='{}',ClientBillingCode='{}',BusinessUnitCode='{}',JobCode='{}',EmploymentStartDate='{}',EmploymentEndDate='{}',WorkAddressFull='{}',WorkLocationCity='{}',WorkLocationState='{}' WHERE BeneficiaryXref='{}' '''.format(WorkLocationCountry,beneficiary_xref, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country,Department,Department_Group,Department_Number,Business_Unit_Code,Client_Billing_Code,ManagerName,ManagerEmail,SecondLevelManager,SecondLevelManagerEmail,BusinessPartnerName,BusinessPartnerEmail,CostCenter,CostCenterNumber,ClientBillingCode,BusinessUnitCode,JobCode,EmploymentStartDate,EmploymentEndDate,WorkAddressFull,WorkLocationCity,WorkLocationState,beneficiary_xref))
                        cursor.commit()

                if beneficiary_xref :
                    if str(beneficiary_type).lower().strip() == 'dependent':

                        primary_ben_data = cursor.execute('''select distinct o.OrganizationXref as Prim_org , p.PetitionerName as primary_petitioner from dbo.Beneficiary b1
                        left join dbo.Beneficiary b2 on b1.PrimaryBeneficiaryXref = b2.BeneficiaryXref
                        left join dbo.Petitioner p on b2.PetitionerXref = p.PetitionerXref 
                        left join dbo.Organization o on b2.OrganizationXref = o.OrganizationXref 
                        where b1.BeneficiaryXref = '{}' '''.format(beneficiary_xref)).fetchone()
                        

                        # using f method is simple n straight forward
                        cursor.execute(f''' update dbo.Beneficiary
                        set OrganizationXref = 
                        '{primary_ben_data.Prim_org}'
                        where BeneficiaryXref = '{beneficiary_xref}' ''')
                        cursor.commit()


                    
             
        # except:
        #     pass

        # except Exception as e :
        #     # print('\n\n\n\nben inconsistent',index, row['primary_beneficiary_id'])      
        #     # print('below exception', e)      
    # time.sleep(3)
    # cursor.execute('''update dbo.Beneficiary
    #         set InactiveDate = NULL  where InactiveDate IN('1900-01-01 00:00:00.000','2021-07-15 00:00:00.000')''')
    # cursor.commit()
            

def process_case_file(file_path, from_name):

    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16
    # df = pd.read_csv(file_path, encoding='latin1',delimiter=',')
    df = pd.read_csv(file_path, encoding = 'unicode_escape', engine ='python',delimiter=',')
    df = df.replace("'","''", regex=True)
    df = df.replace("\.0$","", regex=True)

    # df = df.astype(str)

    # quit()
    
    list_h = df.columns.tolist()
    total_rows = len(df)
   
    for index, row in df.iterrows():

        if True:
            organization_xref = ''
            if 'organization_group_id' in list_h and row['organization_group_id'] and not pd.isna(row['organization_group_id']):
                organization_xref = str(row['organization_group_id']).strip().replace('.0','')
                # organization_xref = fernet.encrypt(organization_xref.encode())

            organization_name = ''
            if 'organization_group_name' in list_h and row['organization_group_name'] and not pd.isna(row['organization_group_name']):
                organization_name = str(str(row['organization_group_name'])).strip()
                if DB_ENCRYPTION == 'YES':
                    organization_name = (fernet.encrypt(organization_name.encode())).decode('utf-8')

            if organization_xref and organization_name:
                results = cursor.execute(
                    '''  SELECT * FROM dbo.Organization where OrganizationXref='{}' '''.format(organization_xref)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute(
                        '''  INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}') '''.format(
                            organization_xref, organization_name))
                    cursor.commit()
                    ##print('inserted')

            petitioner_xref = ''
            if 'petitioner_company_id' in list_h and row['petitioner_company_id'] and not pd.isna(row['petitioner_company_id']):
                petitioner_xref = str(row['petitioner_company_id']).strip().replace('.0','')

            petitioner_name = ''
            if 'petitioner_name' in list_h and row['petitioner_name'] and not pd.isna(row['petitioner_name']):
                petitioner_name = str(str(row['petitioner_name'])).strip()

            beneficiary_xref = ''
            if 'beneficiary_id' in list_h and row['beneficiary_id'] and not pd.isna(row['beneficiary_id']):
                beneficiary_xref = str(row['beneficiary_id']).strip().replace('.0','')

            beneficiary_xref2 = ''
            if 'fn_case_id' in list_h and row['fn_case_id'] and not pd.isna(row['fn_case_id']):
                beneficiary_xref2 = str(row['fn_case_id']).strip().replace('.0','')

            beneficiary_type = ''
            if 'beneficiary_type' in list_h and row['beneficiary_type'] and not pd.isna(row['beneficiary_type']):
                beneficiary_type = str(row['beneficiary_type']).strip()
            
                
            is_primary_beneficiary = 1
            if petitioner_xref and petitioner_name:
                petitioners = cursor.execute('''  select * from dbo.Petitioner where PetitionerXref='{}' '''.format(petitioner_xref)).fetchall()

                if len(petitioners) <= 0 :
                    cursor.execute('''  INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationXref) VALUES ('{}', '{}', '{}') '''.format(petitioner_xref, petitioner_name,organization_xref))
                    cursor.execute('SELECT @@IDENTITY AS ID;')
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()


            if True:
                beneficiary_record_creation_date = ''
                if 'beneficiary_record_opened_date' in list_h and row['beneficiary_record_opened_date'] and not pd.isna(
                        row['beneficiary_record_opened_date']):
                    beneficiary_record_creation_date = change_format(row['beneficiary_record_opened_date'])

                beneficiary_record_inactivation_date = ''
                if 'beneficiary_retired_date' in list_h and row['beneficiary_retired_date'] and not pd.isna(
                        row['beneficiary_retired_date']):
                    beneficiary_record_inactivation_date = change_format(row['beneficiary_retired_date'])

                beneficiary_record_status = 0
                if 'beneficiary_status' in list_h and row['beneficiary_status'] and not pd.isna(row['beneficiary_status']):
                    beneficiary_record_status_chk = str(row['beneficiary_status']).strip()
                    if beneficiary_record_status_chk == 'Active':
                        beneficiary_record_status = 1
                    else:
                        beneficiary_record_status = 0

                beneficiary_last_name = ''
                if 'beneficiary_last_name' in list_h and row['beneficiary_last_name'] and not pd.isna(row['beneficiary_last_name']):
                    beneficiary_last_name = str(str(row['beneficiary_last_name']).strip())

                beneficiary_first_name = ''
                if 'beneficiary_first_name' in list_h and row['beneficiary_first_name'] and not pd.isna(row['beneficiary_first_name']):
                    beneficiary_first_name = str(str(row['beneficiary_first_name']).strip())
                    if DB_ENCRYPTION == 'YES':
                        beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode('utf-8')

                beneficiary_full_name = ''
                if 'full_name' in list_h and row['full_name'] and not pd.isna(row['full_name']):
                    beneficiary_full_name = str(str(row['full_name']).strip())

                primary_beneficiary_id = ''
                if 'primary_fnl_id' in list_h and row['primary_fnl_id'] and not pd.isna(row['primary_fnl_id']):
                    primary_beneficiary_id = str(row['primary_fnl_id']).strip()

                if beneficiary_type == 'Dependent':
                    is_primary_beneficiary = 0
                else:
                    is_primary_beneficiary = 1

                primary_beneficiary_last_name = ''
                if 'primary_last_name' in list_h and row['primary_last_name'] and not pd.isna(row['primary_last_name']):
                    primary_beneficiary_last_name = (row['primary_last_name'].strip())
                    if DB_ENCRYPTION == 'YES':
                        primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode(
                            'utf-8')

                primary_beneficiary_first_name = ''
                if 'primary_first_name' in list_h and row['primary_first_name'] and not pd.isna(row['primary_first_name']):
                    primary_beneficiary_first_name = (row['primary_first_name'].strip())
                    if DB_ENCRYPTION == 'YES':
                        primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode(
                            'utf-8')
                
                primary_beneficiary_full_name = ''
                if 'primary_full_name' in list_h and row['primary_full_name'] and not pd.isna(row['primary_full_name']):
                    primary_beneficiary_full_name = (row['primary_full_name'].strip())
                    if DB_ENCRYPTION == 'YES':
                        primary_beneficiary_full_name = (fernet.encrypt(primary_beneficiary_full_name.encode())).decode(
                            'utf-8')

                relation = ''
                if 'relationship' in list_h and row['relationship'] and not pd.isna(row['relationship']):
                    relation = row['relationship'].strip()
                    if DB_ENCRYPTION == 'YES':
                        relation = (fernet.encrypt(relation.encode())).decode('utf-8')

                immigration_status = ''
                if 'i94_status' in list_h and row['i94_status'] and not pd.isna(row['i94_status']):
                    immigration_status = str(row['i94_status']).strip()

                immigration_status_expiration_status = ''
                if 'i94_expiration' in list_h and row['i94_expiration'] and not pd.isna(row['i94_expiration']):
                    if row['i94_expiration'].strip() == 'D/S':
                        immigration_status_expiration_status = 'D/S'
                    else:
                        if 'D/S' in row['i94_expiration']:
                            split1 = str(str(row['i94_expiration']).strip()).split('(D/S)')
                            immigration_status_expiration_status = change_format(split1[0])
                            immigration_status_expiration_status = str(immigration_status_expiration_status) + ' (D/S)'
                        else:
                            immigration_status_expiration_status = change_format(row['i94_expiration'])

                Current_Immigration_Status = ''
                if 'current_status' in list_h and row['current_status'] and not pd.isna(row['current_status']):
                    Current_Immigration_Status =  str(row['current_status']).strip()
                
                current_status_expiration = ''
                if 'current_status_expiration' in list_h and row['current_status_expiration'] and not pd.isna(row['current_status_expiration']):
                    current_status_expiration = change_format(row['current_status_expiration'])


                # as per shiv inputs data from case file should not be considered for ben file so temp disbling the below insert command

                # if beneficiary_xref:
                #     results = cursor.execute('SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}''.format(beneficiary_xref)).fetchall()
                #     length = len(results)
                #     if length <= 0:
                #         cursor.execute('''  INSERT INTO dbo.Beneficiary(BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, ImmigrationStatus, ImmigrationStatusExpirationDate, from_name, is_primary_beneficiary,Beneficiary_Xref2, FullName, PrimaryBeneficiaryFullName, Current_Immigration_Status,CurrentImmigrationStatusExpirationDate2) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}') '''.format(beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, from_name, is_primary_beneficiary,beneficiary_xref2, beneficiary_full_name, primary_beneficiary_full_name, Current_Immigration_Status, current_status_expiration))
                #         cursor.commit()
                    
                    
                    
                    # else:
                    #     beneficiary_id = results[0].BeneficiaryId
                        #cursor.execute('UPDATE  dbo.Beneficiary SET PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', ImmigrationStatus='{}', ImmigrationStatusExpirationDate='{}', from_name='{}', is_primary_beneficiary='{}',Beneficiary_Xref2='{}',FullName='{}',PrimaryBeneficiaryFullName='{}', Current_Immigration_Status='{}', CurrentImmigrationStatusExpirationDate2='{}'   WHERE BeneficiaryId='{}'  '.format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, from_name, is_primary_beneficiary, beneficiary_xref2, beneficiary_full_name, primary_beneficiary_full_name, Current_Immigration_Status, current_status_expiration, beneficiary_id))
                        #cursor.commit()

                ##print('bid ',beneficiary_id)

                if beneficiary_xref:
                    case_xref = ''
                    if 'process_id' in list_h and not pd.isna(row['process_id']):
                        case_xref = (str(row['process_id']).strip()).replace('.0', '')
                    
                    current_process = ''
                    if 'current_process' in list_h and not pd.isna(row['current_process']):
                        current_process = str(row['current_process']).strip()

                    
                    current_process_name = ''
                    if 'current_process_type' in list_h and not pd.isna(row['current_process_type']):
                        current_process_name = str(row['current_process_type']).strip()
                    
                    case_creation_date = ''
                    if 'process_date_opened' in list_h and row['process_date_opened'] and not pd.isna(row['process_date_opened']):
                        case_creation_date = change_format(row['process_date_opened'])

                    case_petition_name = ''
                    if 'process_type' in list_h and not pd.isna(row['process_type']):
                        case_petition_name = str(str(row['process_type']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_petition_name = (fernet.encrypt(case_petition_name.encode())).decode('utf-8')
                    
                    BeneficiaryPreppedForInterview = ''
                    if 'fn_prepped_for_interview' in list_h and not pd.isna(row['fn_prepped_for_interview']):
                        BeneficiaryPreppedForInterview = str(str(row['fn_prepped_for_interview']).strip())

                    case_type = ''
                    if 'process_type' in list_h and not pd.isna(row['process_type']):
                        case_type = str(str(row['process_type']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_type = (fernet.encrypt(case_type.encode())).decode('utf-8')

                    case_description = ''
                    if 'process_reference' in list_h and not pd.isna(row['process_reference']):
                        case_description = str(str(row['process_reference']).strip())
                        if 'please select' in case_description.lower():
                            case_description = ''
                        if DB_ENCRYPTION == 'YES':
                            case_description = (fernet.encrypt(case_description.encode())).decode('utf-8')
                    
                    case_filed_date = ''
                    if 'filed_date' in list_h and row['filed_date'] and not pd.isna(row['filed_date']):
                        case_filed_date = change_format(row['filed_date'].replace('00:00:00', ''))
                    
                    
                    case_receipt_number = ''
                    if 'receipt_number' in list_h and not pd.isna(row['receipt_number']):
                        case_receipt_number = str(row['receipt_number']).strip()
                        if DB_ENCRYPTION == 'YES':
                            case_receipt_number = (fernet.encrypt(case_receipt_number.encode())).decode('utf-8')

                    case_receipt_status = ''
                    if 'Case Receipt Status' in list_h and not pd.isna(row['Case Receipt Status']):
                        case_receipt_status = str(row['Case Receipt Status']).strip()
                        if DB_ENCRYPTION == 'YES':
                            case_receipt_status = (fernet.encrypt(case_receipt_status.encode())).decode('utf-8')
                    
                    CasePetitionId = ''
                    if 'process_petition_id' in list_h and not pd.isna(row['process_petition_id']):
                        CasePetitionId = str(row['process_petition_id']).strip()

                    FinalAction = ''
                    if 'final_action' in list_h and not pd.isna(row['final_action']):
                        FinalAction = str(row['final_action']).strip()

                    FinalActionDate = ''
                    if 'final_action_date' in list_h and row['final_action_date'] and not pd.isna(row['final_action_date']):
                        FinalActionDate = change_format(row['final_action_date'])


                    rfe_audit_received_date = ''
                    if 'rfe_received' in list_h and row['rfe_received'] and not pd.isna(row['rfe_received']):
                        rfe_audit_received_date = change_format(row['rfe_received'])
                    
                    rfe_audit_due_date = ''
                    if 'rfe_due_date' in list_h and row['rfe_due_date'] and not pd.isna(row['rfe_due_date']):
                        rfe_audit_due_date = change_format(row['rfe_due_date'])
                    
                    approval_of_perm_memo_received = ''
                    if 'approval_of_perm_memo_received' in list_h and row['approval_of_perm_memo_received'] and not pd.isna(row['approval_of_perm_memo_received']):
                        approval_of_perm_memo_received = change_format(row['approval_of_perm_memo_received'])

                    rfe_audit_submitted_date = ''
                    if 'rfe_response_submitted' in list_h and row['rfe_response_submitted'] and not pd.isna(row['rfe_response_submitted']):
                        rfe_audit_submitted_date = change_format(row['rfe_response_submitted'])

                    primary_case_status = ''
                    if 'process_status' in list_h and not pd.isna(row['process_status']):
                        primary_case_status = str(row['process_status']).strip()

                    secondary_case_status = ''
                    if 'final_action' in list_h and not pd.isna(row['final_action']):
                        secondary_case_status = str(str(row['final_action']).strip())
                    
                    case_comments = ''
                    if 'summary_case_disposition' in list_h and not pd.isna(row['summary_case_disposition']):
                        case_comments = str(str(row['summary_case_disposition']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_comments = (fernet.encrypt(case_comments.encode())).decode('utf-8')

                    case_last_step_completed = ''
                    if 'last_process_activity' in list_h and not pd.isna(row['last_process_activity']):
                        case_last_step_completed = str(str(row['last_process_activity']).strip())
                        case_last_step_completed = case_last_step_completed
                        if DB_ENCRYPTION == 'YES':
                            case_last_step_completed = (fernet.encrypt(case_last_step_completed.encode())).decode('utf-8')

                    case_last_step_completed_date = ''
                    if 'last_process_activity_date' in list_h and row['last_process_activity_date'] and not pd.isna(row['last_process_activity_date']):
                        case_last_step_completed_date = change_format(row['last_process_activity_date'])
                    # print('fdsfs',case_last_step_completed_date) 

                    case_next_step_to_be_completed = ''
                    if 'next_unfinished_reminder_subject' in list_h and not pd.isna(row['next_unfinished_reminder_subject']):
                        case_next_step_to_be_completed = str(str(row['next_unfinished_reminder_subject']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_next_step_to_be_completed = (fernet.encrypt(case_next_step_to_be_completed.encode())).decode('utf-8')
                    
                    case_next_step_to_be_completed_date = ''
                    if 'next_unfinished_reminder_expiry' in list_h and row['next_unfinished_reminder_expiry'] and not pd.isna(row['next_unfinished_reminder_expiry']):
                        case_next_step_to_be_completed_date = change_format(row['next_unfinished_reminder_expiry']).replace('00:00:00', '')
                    
                    case_priority_date = ''
                    if 'Case Priority Date' in list_h and row['Case Priority Date'] and not pd.isna(row['Case Priority Date']):
                        case_priority_date = change_format(row['Case Priority Date'])

                    case_priority_category = ''
                    if 'Case Priority Category' in list_h and not pd.isna(row['Case Priority Category']):
                        case_priority_category = str(row['Case Priority Category']).strip()
                        if DB_ENCRYPTION == 'YES':
                            case_priority_category = (fernet.encrypt(case_priority_category.encode())).decode('utf-8')

                    case_priority_country = ''
                    if 'Case Priority Country' in list_h and not pd.isna(row['Case Priority Country']):
                        case_priority_country = str(row['Case Priority Country']).strip()
                        if DB_ENCRYPTION == 'YES':
                            case_priority_country = (fernet.encrypt(case_priority_country.encode())).decode('utf-8')

                    case_approved_date = '' 
                    if 'approval_date' in list_h and row['approval_date'] and not pd.isna(row['approval_date']):
                        case_approved_date = change_format(row['approval_date'])
                    
                    case_valid_from = ''
                    if 'status_valid_from' in list_h and row['status_valid_from'] and not pd.isna(row['status_valid_from']):
                        case_valid_from = change_format(row['status_valid_from'])
                    
                    elif 'date_status_valid_from' in list_h and row['date_status_valid_from'] and not pd.isna(row['date_status_valid_from']):
                        case_valid_from = change_format(row['date_status_valid_from'])
                    
                    case_valid_to = ''
                    if 'status_valid_to' in list_h and row['status_valid_to'] and not pd.isna(row['status_valid_to']):
                        case_valid_to = change_format(row['status_valid_to'])

                    elif 'date_status_valid_to' in list_h and row['date_status_valid_to'] and not pd.isna(row['date_status_valid_to']):
                        case_valid_to = change_format(row['date_status_valid_to'])
                    
                    case_closed_date = ''
                    if 'date_closed' in list_h and row['date_closed'] and not pd.isna(row['date_closed']):
                        case_closed_date = change_format(row['date_closed'])

                    date_labor_certification_expires = ''
                    if 'date_labor_certification_expires' in list_h and row['date_labor_certification_expires'] and not pd.isna(row['date_labor_certification_expires']):
                        date_labor_certification_expires = change_format(row['date_labor_certification_expires'])
                    
                    case_denied_date = ''
                    if 'date_denied' in list_h and row['date_denied'] and not pd.isna(row['date_denied']):
                        case_denied_date = change_format(row['date_denied'])
                    
                    case_withdrawn_date = ''
                    if 'Case Withdrawn Date' in list_h and row['Case Withdrawn Date'] and not pd.isna(row['Case Withdrawn Date']):
                        case_withdrawn_date = change_format(row['Case Withdrawn Date'])
                    
                    case_primary_attorney = ''
                    if 'Case Primary Attorney' in list_h and not pd.isna(row['Case Primary Attorney']):
                        case_primary_attorney = str(str(row['Case Primary Attorney']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_primary_attorney = (fernet.encrypt(case_primary_attorney.encode())).decode('utf-8')
                    
                    case_reviewing_attorney = ''
                    if 'Case Reviewing Attorney' in list_h and not pd.isna(row['Case Reviewing Attorney']):
                        case_reviewing_attorney = str(str(row['Case Reviewing Attorney']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_reviewing_attorney = (fernet.encrypt(case_reviewing_attorney.encode())).decode('utf-8')
                    
                    case_primary_case_manager = ''
                    if 'Case Primary Case Manager' in list_h and not pd.isna(row['Case Primary Case Manager']):
                        case_primary_case_manager = str(str(row['Case Primary Case Manager']).strip())
                        if DB_ENCRYPTION == 'YES':
                            case_primary_case_manager = (fernet.encrypt(case_primary_case_manager.encode())).decode('utf-8')
                    
                    petition_xref = ''
                    if 'process_petition_id' in list_h and not pd.isna(row['process_petition_id']):
                        petition_xref = (str(row['process_petition_id']).strip()).replace('.0', '')


                    ####
                    CaseReceivedDate= ''
                    if 'receipt_date' in list_h and str(row['receipt_date']).strip() and not pd.isna(row['receipt_date']):
                        CaseReceivedDate= change_format(row['receipt_date'])

                    RFEDocsReqestedDate = ''
                    if 'rfe_docs_requested' in list_h and str(row['rfe_docs_requested']).strip() and not pd.isna(row['rfe_docs_requested']):
                        RFEDocsReqestedDate = change_format(row['rfe_docs_requested'])

                    RFEDocsReceivedDate = ''
                    if 'rfe_docs_received' in list_h and str(row['rfe_docs_received']).strip() and not pd.isna(row['rfe_docs_received']):
                        RFEDocsReceivedDate = change_format(row['rfe_docs_received'])

                    PERMAuditReceivedDate = ''
                    if 'audit_notice_received' in list_h and str(row['audit_notice_received']).strip() and not pd.isna(
                            row['audit_notice_received']):
                        PERMAuditReceivedDate = change_format(row['audit_notice_received'])

                    PERMAuditSubmittedDate = ''
                    if 'audit_response_sent_to_dol' in list_h and str(row['audit_response_sent_to_dol']).strip() and not pd.isna(
                            row['audit_response_sent_to_dol']):
                        PERMAuditSubmittedDate = change_format(row['audit_response_sent_to_dol'])

                    SecondaryCaseStatusDate= ''
                    if 'final_action_date' in list_h and str(row['final_action_date']).strip() and not pd.isna(row['final_action_date']):
                        SecondaryCaseStatusDate= change_format(row['final_action_date'])

                    DaysSinceLastStepCompleted= ''
                    if 'days_since_last_activity' in list_h and not pd.isna(row['days_since_last_activity']):
                        DaysSinceLastStepCompleted= (str(row['days_since_last_activity']).strip()).replace('.0', '')

                    visa_preference_category= ''
                    if 'visa_preference_category' in list_h and not pd.isna(row['visa_preference_category']):
                        visa_preference_category= str(row['visa_preference_category']).strip()

                    visa_priority_country= ''
                    if 'visa_priority_country' in list_h and not pd.isna(row['visa_priority_country']):
                        visa_priority_country= str(row['visa_priority_country']).strip()

                    PartnerXref= ''
                    if 'partner_id' in list_h and not pd.isna(row['partner_id']):
                        PartnerXref= str(row['partner_id']).strip()

                    PartnerLastName= ''
                    if 'partner_last_name' in list_h and not pd.isna(row['partner_last_name']):
                        PartnerLastName= str(row['partner_last_name']).strip()

                    PartnerFirstName= ''
                    if 'partner_first_name' in list_h and not pd.isna(row['partner_first_name']):
                        PartnerFirstName= str(row['partner_first_name']).strip()

                    AssociateXref= ''
                    if 'associate_id' in list_h and not pd.isna(row['associate_id']):
                        AssociateXref= str(row['associate_id']).strip()

                    AssociateLastName= ''
                    if 'associate_last_name' in list_h and not pd.isna(row['associate_last_name']):
                        AssociateLastName= str(row['associate_last_name']).strip()

                    AssociateFirstName= ''
                    if 'associate_first_name' in list_h and not pd.isna(row['associate_first_name']):
                        AssociateFirstName= str(row['associate_first_name']).strip()

                    SupervisoryParalegalXref= ''
                    if 'supervisory_paralegal_id' in list_h and not pd.isna(row['supervisory_paralegal_id']):
                        SupervisoryParalegalXref= (str(row['supervisory_paralegal_id']).strip()).replace('.0', '')

                    SupervisoryParalegalLastName= ''
                    if 'supervisory_paralegal_last_name' in list_h and not pd.isna(row['supervisory_paralegal_last_name']):
                        SupervisoryParalegalLastName= str(row['supervisory_paralegal_last_name']).strip()

                    SupervisoryParalegalFirstName= ''
                    if 'supervisory_paralegal_first_name' in list_h and not pd.isna(row['supervisory_paralegal_first_name']):
                        SupervisoryParalegalFirstName= str(row['supervisory_paralegal_first_name']).strip()

                    ParalegalXref= ''
                    if 'paralegal_id' in list_h and not pd.isna(row['paralegal_id']):
                        ParalegalXref= (str(row['paralegal_id']).strip()).replace('.0', '')

                    ParalegalLastName= ''
                    if 'paralegal_last_name' in list_h and not pd.isna(row['paralegal_last_name']):
                        ParalegalLastName= str(row['paralegal_last_name']).strip()

                    ParalegalFirstName= ''
                    if 'paralegal_first_name' in list_h and not pd.isna(row['paralegal_first_name']):
                        ParalegalFirstName= str(row['paralegal_first_name']).strip()

                    AccountManagerXref= ''
                    if 'account_manager_id' in list_h and not pd.isna(row['account_manager_id']):
                        AccountManagerXref= str(row['account_manager_id']).strip()

                    AccountManagerLastName= ''
                    if 'account_manager_last_name' in list_h and not pd.isna(row['account_manager_last_name']):
                        AccountManagerLastName= str(row['account_manager_last_name']).strip()

                    AccountManagerFirstName= ''
                    if 'account_manager_first_name' in list_h and not pd.isna(row['account_manager_first_name']):
                        AccountManagerFirstName= str(row['account_manager_first_name']).strip()

                    SpecialInstructionFlag= ''
                    if 'special_instruction_flag' in list_h and not pd.isna(row['special_instruction_flag']):
                        SpecialInstructionFlag= str(row['special_instruction_flag']).strip()

                    # SpecialInstructionInfo= ''
                    # if 'special_instruction_info' in list_h and not pd.isna(row['special_instruction_info']):
                    #     SpecialInstructionInfo= str(row['special_instruction_info']).strip()

                    SpecialInstructionInfo= ''
                    if 'hr_instruction_flag' in list_h and not pd.isna(row['hr_instruction_flag']):
                        SpecialInstructionInfo= str(row['hr_instruction_flag']).strip()


                    ClientBillingCode= ''
                    if 'client_billing_code' in list_h and not pd.isna(row['client_billing_code']):
                        ClientBillingCode= str(row['client_billing_code']).strip()

                    OnlineIntakeDate= ''
                    if 'online_intake_date' in list_h and str(row['online_intake_date']).strip() and not pd.isna(row['online_intake_date']):
                        OnlineIntakeDate= change_format(row['online_intake_date'])


                    questionnairesenttomanager= ''
                    if 'questionnaire_sent_to_manager' in list_h and not pd.isna(row['questionnaire_sent_to_manager']):
                        questionnairesenttomanager= change_format(str(row['questionnaire_sent_to_manager']).strip())

                    questionnairessenttofn= ''
                    if 'questionnaires_sent_to_fn' in list_h and not pd.isna(row['questionnaires_sent_to_fn']):
                        questionnairessenttofn= change_format(str(row['questionnaires_sent_to_fn']).strip())

                    followupwithfnforrequestedinformation= ''
                    if 'follow_up_with_fn_for_requested_information' in list_h and not pd.isna(row['follow_up_with_fn_for_requested_information']):
                        followupwithfnforrequestedinformation= change_format(str(row['follow_up_with_fn_for_requested_information']).strip())

                    questionnairecompletedandreturnedbymanager= ''
                    if 'questionnaire_completed_and_returned_by_manager' in list_h and not pd.isna(row['questionnaire_completed_and_returned_by_manager']):
                        questionnairecompletedandreturnedbymanager= change_format(str(row['questionnaire_completed_and_returned_by_manager']).strip())

                    questionnairecompletedandreturnedbyfn= ''
                    if 'questionnaire_completed_and_returned_by_fn' in list_h and not pd.isna(row['questionnaire_completed_and_returned_by_fn']):
                        questionnairecompletedandreturnedbyfn= change_format(str(row['questionnaire_completed_and_returned_by_fn']).strip())

                    employersubmissionquestionnairecompleted= ''
                    if 'employer_submission_questionnaire_completed' in list_h and not pd.isna(row['employer_submission_questionnaire_completed']):
                        employersubmissionquestionnairecompleted= change_format(str(row['employer_submission_questionnaire_completed']).strip())

                    allpetitioningcompanyinforeceived= ''
                    if 'all_petitioning_company_info_received' in list_h and not pd.isna(row['all_petitioning_company_info_received']):
                        allpetitioningcompanyinforeceived= change_format(str(row['all_petitioning_company_info_received']).strip())

                    allfndocsreceived= ''
                    if 'all_fn_docs_received' in list_h and not pd.isna(row['all_fn_docs_received']):
                        allfndocsreceived= change_format(str(row['all_fn_docs_received']).strip())

                    fncompletedquestionnairesandacknowledgement= ''
                    if 'fn_completed_questionnaires_and_acknowledgement' in list_h and not pd.isna(row['fn_completed_questionnaires_and_acknowledgement']):
                        fncompletedquestionnairesandacknowledgement= change_format(str(row['fn_completed_questionnaires_and_acknowledgement']).strip())

                    fnquestionnairescompleted= ''
                    if 'fn_questionnaires_completed' in list_h and not pd.isna(row['fn_questionnaires_completed']):
                        fnquestionnairescompleted= change_format(str(row['fn_questionnaires_completed']).strip())

                    lcafiled= ''
                    if 'lca_filed' in list_h and not pd.isna(row['lca_filed']):
                        lcafiled= change_format(str(row['lca_filed']).strip())

                    lcacasenumber= ''
                    if 'lca_case_number' in list_h and not pd.isna(row['lca_case_number']):
                        lcacasenumber= str(row['lca_case_number']).strip()

                    lcacertified= ''
                    if 'lca_certified' in list_h and not pd.isna(row['lca_certified']):
                        lcacertified= change_format(str(row['lca_certified']).strip())

                    formsanddocumentationprepped= ''
                    if 'forms_and_documentation_prepped' in list_h and not pd.isna(row['forms_and_documentation_prepped']):
                        formsanddocumentationprepped= change_format(str(row['forms_and_documentation_prepped']).strip())

                    formsanddocumentationsubmittedforsignature= ''
                    if 'forms_and_documentation_submitted_for_signature' in list_h and not pd.isna(row['forms_and_documentation_submitted_for_signature']):
                        formsanddocumentationsubmittedforsignature= change_format(str(row['forms_and_documentation_submitted_for_signature']).strip())

                    signedformsandletterreceived= ''
                    if 'signed_forms_and_letter_received' in list_h and not pd.isna(row['signed_forms_and_letter_received']):
                        signedformsandletterreceived= change_format(str(row['signed_forms_and_letter_received']).strip())

                    dateaosformssentforsignature= ''
                    if 'date_aos_forms_sent_for_signature' in list_h and str(row['date_aos_forms_sent_for_signature']).strip() and not pd.isna(row['date_aos_forms_sent_for_signature']):
                        dateaosformssentforsignature= change_format(row['date_aos_forms_sent_for_signature'])

                    datesignedaosformsreceived= ''
                    if 'date_signed_aos_forms_received' in list_h and str(row['date_signed_aos_forms_received']).strip() and not pd.isna(row['date_signed_aos_forms_received']):
                        datesignedaosformsreceived= change_format(row['date_signed_aos_forms_received'])

                    targetfiledate= ''
                    if 'target_file_date' in list_h and str(row['target_file_date']).strip() and not pd.isna(row['target_file_date']):
                        targetfiledate= change_format((row['target_file_date']))
                        # quit()
                    applicationfiled= ''
                    if 'application_filed' in list_h:
                        applicationfiled= change_format(str(row['application_filed']).strip())

                    applicationfiledwithcis= ''
                    if 'application_filed_with_cis' in list_h:
                        applicationfiledwithcis= change_format(str(row['application_filed_with_cis']).strip())

                    petitionfiledwithcis= ''
                    if 'petition_filed_with_cis' in list_h:
                        petitionfiledwithcis= change_format(str(row['petition_filed_with_cis']).strip())

                    formi129filedwithcis= ''
                    if 'form_i129_filed_with_cis' in list_h:
                        formi129filedwithcis= change_format(str(row['form_i129_filed_with_cis']).strip())
                    
                    aosapplicationfiled = ''
                    if 'aos_application_filed' in list_h:
                        aosapplicationfiled = change_format(str(row['aos_application_filed']).strip())

                    tnpacketsenttofnforpoeprocessing= ''
                    if 'tn_packet_sent_to_fn_for_poe_processing' in list_h:
                        tnpacketsenttofnforpoeprocessing = change_format(str(row['tn_packet_sent_to_fn_for_poe_processing']).strip())

                    appealmotionduedate= ''
                    if 'appeal_motion_due_date' in list_h and str(row['appeal_motion_due_date']).strip() and not pd.isna(row['appeal_motion_due_date']):
                        appealmotionduedate= change_format(row['appeal_motion_due_date'])

                    appealmotionfiled= ''
                    if 'appeal_motion_filed' in list_h:
                        appealmotionfiled= change_format(str(row['appeal_motion_filed']).strip())

                    consularinterviewdate = ''
                    if 'consular_interview_date' in list_h and str(row['consular_interview_date']).strip() and not pd.isna(
                            row['consular_interview_date']):
                        consularinterviewdate = change_format(row['consular_interview_date'])

                    supplementalbriefdocsfiled= ''
                    if 'supplemental_brief_docs_filed' in list_h:
                        supplementalbriefdocsfiled= change_format(str(row['supplemental_brief_docs_filed']).strip())

                    docketdatebalca = ''
                    if 'docket_date_balca' in list_h and str(row['docket_date_balca']).strip() and not pd.isna(row['docket_date_balca']):
                        docketdatebalca = change_format(row['docket_date_balca'])

                    datewithdrawrequestsenttouscis = ''
                    if 'date_withdraw_request_sent_to_uscis' in list_h and str(row['date_withdraw_request_sent_to_uscis']).strip() and not pd.isna(
                            row['date_withdraw_request_sent_to_uscis']):
                        datewithdrawrequestsenttouscis = change_format(row  ['date_withdraw_request_sent_to_uscis'])

                    withdrawalrequestconfirmedbydoluscis= ''
                    if 'withdrawal_request_confirmed_by_dol_uscis' in list_h:
                        withdrawalrequestconfirmedbydoluscis= change_format(str(row['withdrawal_request_confirmed_by_dol_uscis']).strip())

                    approvalpackagesent= ''
                    if 'approval_package_sent' in list_h:
                        approvalpackagesent= change_format(str(row['approval_package_sent']).strip())

                    h1bregistrationsubmitted= ''
                    if 'h1b_registration_submitted' in list_h:
                        h1bregistrationsubmitted= change_format(str(row['h1b_registration_submitted']).strip())

                    h1bregistrationresult= ''
                    if 'h1b_registration_result' in list_h:
                        h1bregistrationresult= change_format(str(row['h1b_registration_result']).strip())

                    h1bcapregistrationselected= ''
                    if 'h1b_cap_registration_selected' in list_h:
                        h1bcapregistrationselected= change_format(str(row['h1b_cap_registration_selected']).strip())

                    I907filedupgradedtopremprocessing= ''
                    if 'i907_filed_upgraded_to_prem_processing' in list_h:
                        I907filedupgradedtopremprocessing= change_format(str(row['i907_filed_upgraded_to_prem_processing']).strip())

                    premiumprocessingfeereceivedfromfn= ''
                    if 'premium_processing_fee_received_from_fn' in list_h:
                        premiumprocessingfeereceivedfromfn= change_format(str(row['premium_processing_fee_received_from_fn']).strip())

                    receipts= ''
                    if 'receipts' in list_h and not pd.isna(row['receipts']):
                        receipts= str(row['receipts']).strip()

                    I485receiptdate= ''
                    if 'i485_receipt_date' in list_h and str(row['i485_receipt_date']).strip() and not pd.isna(row['i485_receipt_date']):
                        I485receiptdate= change_format(row['i485_receipt_date'])

                    I485jportabilityreceiptdate= ''
                    if 'i485j_portability_receipt_date' in list_h and str(row['i485j_portability_receipt_date']).strip() and not pd.isna(row['i485j_portability_receipt_date']):
                        I485jportabilityreceiptdate= change_format(row['i485j_portability_receipt_date'])

                    I131receiptdate= ''
                    if 'i131_receipt_date' in list_h and str(row['i131_receipt_date']).strip() and not pd.isna(row['i131_receipt_date']):
                        I131receiptdate= change_format(row['i131_receipt_date'])
                    
                    PwdExpirationDate= ''
                    if 'pwd_expiration_date' in list_h and str(row['pwd_expiration_date']).strip() and not pd.isna(row['pwd_expiration_date']):
                        PwdExpirationDate = change_format(row['pwd_expiration_date'])


                    apreceiptnoticereceived= ''
                    if 'ap_receipt_notice_received' in list_h:
                        apreceiptnoticereceived= change_format(str(row['ap_receipt_notice_received']).strip())

                    eadreceiptnoticereceived= ''
                    if 'ead_receipt_notice_received' in list_h:
                        eadreceiptnoticereceived= change_format(str(row['ead_receipt_notice_received']).strip())

                    petitioningjobtitle= ''
                    if 'petitioning_job_title' in list_h and not pd.isna(row['petitioning_job_title']):
                        petitioningjobtitle= str(row['petitioning_job_title']).strip()

                    petitioningjoblocation= ''
                    if 'petitioning_job_location' in list_h and not pd.isna(row['petitioning_job_location']):
                        petitioningjoblocation= str(row['petitioning_job_location']).strip()

                    permmemosenttoemployer= ''
                    if 'perm_memo_sent_to_employer' in list_h:
                        permmemosenttoemployer= change_format(str(row['perm_memo_sent_to_employer']).strip())

                    approvalofpermmemoreceived= ''
                    if 'approvalofpermmemoreceived' in list_h and row['approvalofpermmemoreceived'] and not pd.isna(row['approvalofpermmemoreceived']):
                        approvalofpermmemoreceived= change_format(str(row['approvalofpermmemoreceived']).strip())

                    elif 'approval_of_perm_memo_received' in list_h and row['approval_of_perm_memo_received'] and not pd.isna(row['approval_of_perm_memo_received']):
                        approvalofpermmemoreceived= change_format(str(row['approval_of_perm_memo_received']).strip())

                    employeeworkexperiencechartsent= ''
                    if 'employee_work_experience_chart_sent' in list_h:
                        employeeworkexperiencechartsent= change_format(str(row['employee_work_experience_chart_sent']).strip())

                    employeeworkexperiencechartreceived= ''
                    if 'employee_work_experience_chart_received' in list_h:
                        employeeworkexperiencechartreceived = change_format(str(row['employee_work_experience_chart_received']).strip())

                    employmentverificationletterssenttoemployee= ''
                    if 'employment_verification_letters_sent_to_employee' in list_h:
                        employmentverificationletterssenttoemployee= change_format(str(row['employment_verification_letters_sent_to_employee']).strip())

                    signedemploymentverificationlettersreceived= ''
                    if 'signed_employment_verification_letters_received' in list_h:
                        signedemploymentverificationlettersreceived= change_format(str(row['signed_employment_verification_letters_received']).strip())

                    prevailingwagedeterminationrequestsubmittedtodol= ''
                    if 'prevailing_wage_determination_request_submitted_to_dol' in list_h:
                        prevailingwagedeterminationrequestsubmittedtodol = change_format(str(row['prevailing_wage_determination_request_submitted_to_dol']).strip())

                    prevailingwagedeterminationissuedbydol= ''
                    if 'prevailing_wage_determination_issued_by_dol' in list_h:
                        prevailingwagedeterminationissuedbydol= change_format(str(row['prevailing_wage_determination_issued_by_dol']).strip())

                    recruitmentinstructionssenttocompany= ''
                    if 'recruitment_instructions_sent_to_company' in list_h:
                        recruitmentinstructionssenttocompany= change_format(str(row['recruitment_instructions_sent_to_company']).strip())

                    joborderplacedwithswa= ''
                    if 'job_order_placed_with_swa' in list_h:
                        joborderplacedwithswa= change_format(str(row['job_order_placed_with_swa']).strip())

                    noticeoffilingposted= ''
                    if 'notice_of_filing_posted' in list_h:
                        noticeoffilingposted= change_format(str(row['notice_of_filing_posted']).strip())

                    intranetnoticeoffilingposted= ''
                    if 'intranet_notice_of_filing_posted' in list_h:
                        intranetnoticeoffilingposted= change_format(str(row['intranet_notice_of_filing_posted']).strip())

                    noticeoffilingremovedsigned= ''
                    if 'notice_of_filing_removed_signed' in list_h:
                        noticeoffilingremovedsigned= change_format(str(row['notice_of_filing_removed_signed']).strip())

                    intranetnoticeoffilingremoved= ''
                    if 'intranet_notice_of_filing_removed' in list_h:
                        intranetnoticeoffilingremoved= change_format(str(row['intranet_notice_of_filing_removed']).strip())

                    _1stsundayadplaced= ''
                    if '_1st_sunday_ad_placed' in list_h:
                        _1stsundayadplaced= change_format(str(row['_1st_sunday_ad_placed']).strip())

                    _2ndsundayadplaced= ''
                    if '_2nd_sunday_ad_placed' in list_h:
                        _2ndsundayadplaced= change_format(str(row['_2nd_sunday_ad_placed']).strip())

                    _1stadditionalrecruitmentstepplaced= ''
                    if '_1st_additional_recruitment_step_placed' in list_h:
                        _1stadditionalrecruitmentstepplaced = change_format(str(row['_1st_additional_recruitment_step_placed']).strip())

                    _2ndadditionalrecruitmentstepplaced= ''
                    if '_2nd_additional_recruitment_step_placed' in list_h:
                        _2ndadditionalrecruitmentstepplaced= change_format(str(row['_2nd_additional_recruitment_step_placed']).strip())

                    _3rdadditionalrecruitmentstepplaced= ''
                    if '_3rd_additional_recruitment_step_placed' in list_h:
                        _3rdadditionalrecruitmentstepplaced= change_format(str(row['_3rd_additional_recruitment_step_placed']).strip())

                    datedcopiesofallrecruitmentreceived= ''
                    if 'dated_copies_of_all_recruitment_received' in list_h:
                        datedcopiesofallrecruitmentreceived = change_format(str(row['dated_copies_of_all_recruitment_received']).strip())

                    completedevaluationquestionnairesandresumesreceived= ''
                    if 'completed_evaluation_questionnaires_and_resumes_received' in list_h:
                        completedevaluationquestionnairesandresumesreceived = change_format(str(row['completed_evaluation_questionnaires_and_resumes_received']).strip())

                    recruitmentreportsenttocompany= ''
                    if 'recruitment_report_sent_to_company' in list_h:
                        recruitmentreportsenttocompany= change_format(str(row['recruitment_report_sent_to_company']).strip())

                    recruitmentreportreceived= ''
                    if 'recruitment_report_received' in list_h:
                        recruitmentreportreceived= change_format(str(row['recruitment_report_received']).strip())

                    form9089senttofnandemployer= ''
                    if 'form_9089_sent_to_fn_and_employer' in list_h:
                        form9089senttofnandemployer= change_format(str(row['form_9089_sent_to_fn_and_employer']).strip())

                    editstoform9089receivedfromfnandemployer= ''
                    if 'edits_to_form_9089_received_from_fn_and_employer' in list_h:
                        editstoform9089receivedfromfnandemployer= change_format(str(row['edits_to_form_9089_received_from_fn_and_employer']).strip())

                    form9089submittedtodol= ''
                    if 'form_9089_submitted_to_dol' in list_h:
                        form9089submittedtodol= change_format(str(row['form_9089_submitted_to_dol']).strip())

                    inputcallconducted= ''
                    if 'input_call_conducted' in list_h:
                        inputcallconducted= change_format(str(row['input_call_conducted']).strip())

                    inputstatementreceived= ''
                    if 'input_statement_received' in list_h:
                        inputstatementreceived= change_format(str(row['input_statement_received']).strip())

                    casestrategyandlettersplansent= ''
                    if 'case_strategy_and_letters_plan_sent' in list_h:
                        casestrategyandlettersplansent= change_format(str(row['case_strategy_and_letters_plan_sent']).strip())

                    longlettersenttofn= ''
                    if 'long_letter_sent_to_fn' in list_h:
                        longlettersenttofn= change_format(str(row['long_letter_sent_to_fn']).strip())

                    shortletterssenttofn= ''
                    if 'short_letters_sent_to_fn' in list_h:
                        shortletterssenttofn= change_format(str(row['short_letters_sent_to_fn']).strip())

                    numberoftotalapplicants= ''
                    if 'number_of_total_applicants' in list_h:
                        numberoftotalapplicants= change_format(str(row['number_of_total_applicants']).strip())

                    numberofnonusworkers= ''
                    if 'number_of_non_us_workers' in list_h:
                        numberofnonusworkers= change_format(str(row['number_of_non_us_workers']).strip())

                    numberofphonescreensconducted= ''
                    if 'number_of_phone_screens_conducted' in list_h:
                        numberofphonescreensconducted= change_format(str(row['number_of_phone_screens_conducted']).strip())

                    numberofmanagerinterviewsconducted= ''
                    if 'number_of_manager_interviews_conducted' in list_h:
                        numberofmanagerinterviewsconducted= change_format(str(row['number_of_manager_interviews_conducted']).strip())



                    ##print('cx ', case_xref)
                    if case_xref:
                        
                        ##print('SELECT * FROM [dbo].[Case] where BeneficiaryId='{}' and CaseXref='{}' and from_name='{}''.format(beneficiary_id, case_xref, from_name))
                        results = cursor.execute('''   SELECT * FROM [dbo].[Case] where CaseXref='{}' '''.format(case_xref)).fetchall()
                        length = len(results)
                        if length <= 0:
                            cursor.execute('''  INSERT INTO [dbo].[Case](FinalActionDate,BeneficiaryPreppedForInterview,CaseReceivedDate,CasePetitionId,FinalAction,OrganizationXref,CaseXref, BeneficiaryXref, SourceCreatedDate, CasePetitionName, CaseType, CaseDescription, CaseFiledDate, ReceiptNumber, ReceiptStatus, RFEAuditReceivedDate,RFEAuditDueDate, RFEAuditSubmittedDate, PrimaryCaseStatus, SecondaryCaseStatus, CaseComments, LastStepCompleted, LastStepCompletedDate, NextStepAction, NextStepActionDueDate, PriorityDate, PriorityCategory, PriorityCountry, CaseApprovedDate, CaseValidFromDate, CaseExpirationDate, CaseClosedDate, CaseDeniedDate, CaseWithdrawnDate, CasePrimaryAttorney, CaseReviewingAttorney, CasePrimaryCaseManager, PetitionXref, IsCurrentProcess, CurrentProcessName, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing, appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted,Form9089expirationdate,PwdExpirationDate) VALUES ('{}','{}','{}','{}','{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}','{}','{}') '''.format(FinalActionDate,BeneficiaryPreppedForInterview,CaseReceivedDate,CasePetitionId,FinalAction,organization_xref,case_xref, beneficiary_xref, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref,  current_process, current_process_name, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing,appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted,date_labor_certification_expires,PwdExpirationDate))
                            cursor.commit()
                        else:
                            cursor.execute('''  UPDATE [dbo].[Case] SET FinalActionDate = '{}',BeneficiaryPreppedForInterview = '{}',CaseReceivedDate = '{}', CasePetitionId = '{}', FinalAction = '{}',PwdExpirationDate='{}',OrganizationXref='{}',CaseXref='{}', BeneficiaryXref='{}', SourceCreatedDate='{}', CasePetitionName='{}', CaseType='{}', CaseDescription='{}', CaseFiledDate='{}', ReceiptNumber='{}', ReceiptStatus='{}', RFEAuditReceivedDate='{}', RFEAuditDueDate='{}', RFEAuditSubmittedDate='{}', PrimaryCaseStatus='{}', SecondaryCaseStatus='{}', CaseComments='{}', LastStepCompleted='{}', LastStepCompletedDate='{}', NextStepAction='{}', NextStepActionDueDate='{}', PriorityDate='{}', PriorityCategory='{}', PriorityCountry='{}', CaseApprovedDate='{}', CaseValidFromDate='{}', CaseExpirationDate='{}', CaseClosedDate='{}', CaseDeniedDate='{}', CaseWithdrawnDate='{}', CasePrimaryAttorney='{}', CaseReviewingAttorney='{}', CasePrimaryCaseManager='{}', PetitionXref='{}',IsCurrentProcess='{}', CurrentProcessName='{}' ,RFEDocsReqestedDate='{}',  RFEDocsReceivedDate='{}',  SecondaryCaseStatusDate='{}',  DaysSinceLastStepCompleted='{}',  visa_preference_category='{}',  visa_priority_country='{}',  PartnerXref='{}',  PartnerLastName='{}',  PartnerFirstName='{}',  AssociateXref='{}',  AssociateLastName='{}',  AssociateFirstName='{}',  SupervisoryParalegalXref='{}',  SupervisoryParalegalLastName='{}',  SupervisoryParalegalFirstName='{}',  ParalegalXref='{}',  ParalegalLastName='{}',  ParalegalFirstName='{}',  AccountManagerXref='{}',  AccountManagerLastName='{}',  AccountManagerFirstName='{}',  SpecialInstructionFlag='{}',  SpecialInstructionInfo='{}',  ClientBillingCode='{}',  OnlineIntakeDate='{}',  questionnairesenttomanager='{}',  questionnairessenttofn='{}',  followupwithfnforrequestedinformation='{}',  questionnairecompletedandreturnedbymanager='{}',  questionnairecompletedandreturnedbyfn='{}',  employersubmissionquestionnairecompleted='{}',  allpetitioningcompanyinforeceived='{}',  allfndocsreceived='{}',  fncompletedquestionnairesandacknowledgement='{}',  fnquestionnairescompleted='{}',  lcafiled='{}',  lcacasenumber='{}',  lcacertified='{}',  formsanddocumentationprepped='{}',  formsanddocumentationsubmittedforsignature='{}',  signedformsandletterreceived='{}',  dateaosformssentforsignature='{}',  datesignedaosformsreceived='{}',  targetfiledate='{}',  applicationfiled='{}',  applicationfiledwithcis='{}',  petitionfiledwithcis='{}',  formi129filedwithcis='{}',  aosapplicationfiled='{}',  tnpacketsenttofnforpoeprocessing='{}',  appealmotionduedate='{}',  appealmotionfiled='{}',  consularinterviewdate='{}',  supplementalbriefdocsfiled='{}',  docketdatebalca='{}',  datewithdrawrequestsenttouscis='{}',  withdrawalrequestconfirmedbydoluscis='{}',  approvalpackagesent='{}',  h1bregistrationsubmitted='{}',  h1bregistrationresult='{}',  h1bcapregistrationselected='{}',  I907filedupgradedtopremprocessing='{}',  premiumprocessingfeereceivedfromfn='{}',  receipts='{}',  I485receiptdate='{}',  I485jportabilityreceiptdate='{}',  I131receiptdate='{}',  apreceiptnoticereceived='{}',  eadreceiptnoticereceived='{}',  petitioningjobtitle='{}',  petitioningjoblocation='{}',  permmemosenttoemployer='{}',  approvalofpermmemoreceived='{}',  employeeworkexperiencechartsent='{}',  employeeworkexperiencechartreceived='{}',  employmentverificationletterssenttoemployee='{}',  signedemploymentverificationlettersreceived='{}',  prevailingwagedeterminationrequestsubmittedtodol='{}',  prevailingwagedeterminationissuedbydol='{}',  recruitmentinstructionssenttocompany='{}',  joborderplacedwithswa='{}',  noticeoffilingposted='{}',  intranetnoticeoffilingposted='{}',  noticeoffilingremovedsigned='{}',  intranetnoticeoffilingremoved='{}',  _1stsundayadplaced='{}',  _2ndsundayadplaced='{}',  _1stadditionalrecruitmentstepplaced='{}',  _2ndadditionalrecruitmentstepplaced='{}',  _3rdadditionalrecruitmentstepplaced='{}',  datedcopiesofallrecruitmentreceived='{}',  completedevaluationquestionnairesandresumesreceived='{}',  recruitmentreportsenttocompany='{}',  recruitmentreportreceived='{}',  form9089senttofnandemployer='{}',  editstoform9089receivedfromfnandemployer='{}',  form9089submittedtodol='{}',  inputcallconducted='{}',  inputstatementreceived='{}',  casestrategyandlettersplansent='{}',  longlettersenttofn='{}',  shortletterssenttofn='{}',  numberoftotalapplicants='{}',  numberofnonusworkers='{}',  numberofphonescreensconducted='{}',  numberofmanagerinterviewsconducted='{}',Form9089expirationdate='{}' WHERE CaseXref = '{}' '''.format(FinalActionDate,BeneficiaryPreppedForInterview,CaseReceivedDate,CasePetitionId,FinalAction,PwdExpirationDate,organization_xref,case_xref, beneficiary_xref, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref,  current_process, current_process_name, RFEDocsReqestedDate, RFEDocsReceivedDate, SecondaryCaseStatusDate, DaysSinceLastStepCompleted, visa_preference_category, visa_priority_country, PartnerXref, PartnerLastName, PartnerFirstName, AssociateXref, AssociateLastName, AssociateFirstName, SupervisoryParalegalXref, SupervisoryParalegalLastName, SupervisoryParalegalFirstName, ParalegalXref, ParalegalLastName, ParalegalFirstName, AccountManagerXref, AccountManagerLastName, AccountManagerFirstName, SpecialInstructionFlag, SpecialInstructionInfo, ClientBillingCode, OnlineIntakeDate, questionnairesenttomanager, questionnairessenttofn, followupwithfnforrequestedinformation, questionnairecompletedandreturnedbymanager, questionnairecompletedandreturnedbyfn, employersubmissionquestionnairecompleted, allpetitioningcompanyinforeceived, allfndocsreceived, fncompletedquestionnairesandacknowledgement, fnquestionnairescompleted, lcafiled, lcacasenumber, lcacertified, formsanddocumentationprepped, formsanddocumentationsubmittedforsignature, signedformsandletterreceived, dateaosformssentforsignature, datesignedaosformsreceived, targetfiledate, applicationfiled, applicationfiledwithcis, petitionfiledwithcis, formi129filedwithcis, aosapplicationfiled, tnpacketsenttofnforpoeprocessing,appealmotionduedate, appealmotionfiled, consularinterviewdate, supplementalbriefdocsfiled, docketdatebalca, datewithdrawrequestsenttouscis, withdrawalrequestconfirmedbydoluscis, approvalpackagesent, h1bregistrationsubmitted, h1bregistrationresult, h1bcapregistrationselected, I907filedupgradedtopremprocessing, premiumprocessingfeereceivedfromfn, receipts, I485receiptdate, I485jportabilityreceiptdate, I131receiptdate, apreceiptnoticereceived, eadreceiptnoticereceived, petitioningjobtitle, petitioningjoblocation, permmemosenttoemployer, approvalofpermmemoreceived, employeeworkexperiencechartsent, employeeworkexperiencechartreceived, employmentverificationletterssenttoemployee, signedemploymentverificationlettersreceived, prevailingwagedeterminationrequestsubmittedtodol, prevailingwagedeterminationissuedbydol, recruitmentinstructionssenttocompany, joborderplacedwithswa, noticeoffilingposted, intranetnoticeoffilingposted, noticeoffilingremovedsigned, intranetnoticeoffilingremoved, _1stsundayadplaced, _2ndsundayadplaced, _1stadditionalrecruitmentstepplaced, _2ndadditionalrecruitmentstepplaced, _3rdadditionalrecruitmentstepplaced, datedcopiesofallrecruitmentreceived, completedevaluationquestionnairesandresumesreceived, recruitmentreportsenttocompany, recruitmentreportreceived, form9089senttofnandemployer, editstoform9089receivedfromfnandemployer, form9089submittedtodol, inputcallconducted, inputstatementreceived, casestrategyandlettersplansent, longlettersenttofn, shortletterssenttofn, numberoftotalapplicants, numberofnonusworkers, numberofphonescreensconducted, numberofmanagerinterviewsconducted,date_labor_certification_expires,case_xref))
                            cursor.commit()
      
        # except Exception as e:
        #     # print('case inconsistent',index,row)      
        #     # print('\n below exception', e)   
        #     pass 
        
#updation in db table for petitioner details

    cursor.execute(''' 
    update dbo.Beneficiary
    set PrimaryBeneficiaryXref = BeneficiaryXref, RelationType = 'Self'
    where BeneficiaryType = 'Primary'

    update Beneficiary 
    set PetitionerofPrimaryBeneficiary = p.PetitionerName
    FROM Beneficiary b
    join Petitioner p on p.PetitionerXref = b.PetitionerXref
    where b.BeneficiaryType = 'Primary';

    ''')
    cursor.commit()


def changes_to_db_tables():

    tables = ['dbo.Beneficiary','dbo.BeneficiaryPriorityDate','dbo.BeneficiaryPriorityDate','dbo.BeneficiaryEmployment','dbo.[case]']
    for  table in tables:
        table_columns = cursor.execute(''' 
        SELECT name FROM sys.columns WHERE object_id = OBJECT_ID('{}')
        '''.format(table)).fetchall()
        

        #this below loop replaces .0 i numbers
        for i in range(len(table_columns)):     
            try:      
                cursor.execute('''  UPDATE {}
                SET {} = REPLACE({}, '.0','') '''.format(table,table_columns[i].name,table_columns[i].name))
                cursor.commit()
            except:
                pass

        #this below loop replaces nan as NULL 
        for i in range(len(table_columns)):     
            try:      
                cursor.execute('''  update {}
                            set {} = NULL  where {} IN('nan')'''.format(table,table_columns[i].name,table_columns[i].name))
                cursor.commit()
            except:
                pass
    
        #this below loop replaces improper date value as  NULL 
        for i in range(len(table_columns)):
            try:
                cursor.execute('''  update {}
                        set {} = NULL  where {} IN('1900-01-01 00:00:00.000')'''.format(table,table_columns[i].name,table_columns[i].name))
                cursor.commit()
            except:
                pass



def start():
    truncate_full_db()
    # quit()
    current_time = datetime.now() 
    month = str(current_time.month).rjust(2, '0')
    day = str(current_time.day).rjust(2, '0')
    todate = month+''+day+''+str(current_time.year)
    from_name = ''

    if os.path.exists('Source Data/beneficiary.csv'):
        ben_file = True
        print('Processing - beneficiary.csv file')
        process_beneficiary_file('Source Data/beneficiary.csv',from_name)
        print('Pushed Beneficiary Data to DB Successfully..')
    else:
        print('beneficiary.csv file not found in source directory..\nProgram Terminated..')
        quit()
        
    if os.path.exists('Source Data/process.csv'):
        print('Processing - process.csv file')
        process_case_file('Source Data/process.csv',from_name)
        print('Pushed process Data to DB Successfully..')
    else:
        print('process.csv file not found in source directory..\nProgram Terminated..')
        quit()

    # quit()
    changes_to_db_tables()
    if SECURE_DB_TABLES == 'ON':
        secure_tables()
    # quit()
                
    # result_filepath = 'Processed Reports/Charter – Active Employee Report_'+str(todate)+'.xlsx'
    # ActiveEmployeeReport(result_filepath)
    # print('Processed - Charter – Active Employee Report')

    # result_filepath2 = 'Processed Reports/Charter – PERM Report_'+str(todate)+'.xlsx'
    # PermReport(result_filepath2)
    # print('Processed - Charter – PERM Report')

    # result_filepath3 = 'Processed Reports/Charter – Weekly NetOps - P&T Transfer Report_'+str(todate)+'.xlsx'
    # WeeklyNetOps(result_filepath3)
    # print('Processed - Charter – Weekly NetOps - P&T Transfer Report')

    # result_filepath4 = 'Processed Reports/KILP Internal Report - Paralegal Active List_'+str(todate)+'.xlsx'
    # Paralegal(result_filepath4)
    # print('Processed - KILP Internal Report - Paralegal Active List')

    # result_filepath5 = 'Processed Reports/KILP Internal Report - PERM Report_'+str(todate)+'.xlsx'
    # InternalPerm(result_filepath5)
    # print('Processed - KILP Internal Report - PERM Report')

    # result_filepath6 = 'Processed Reports/Comcast Status Report_'+str(todate)+'.xlsx'
    # ComcastReport(result_filepath6)
    # print('Processed - Comcast Status Report')


    # result_filepath7 = 'Processed Reports/Charter Report - Active Beneficiary Report_'+str(todate)+'.xlsx'
    # CharterActiveBenReport(result_filepath7)
    # print('Processed - Charter Report - Active Beneficiary Report')
   
    # result_filepath8 = 'Processed Reports/Charter New Hire Report_'+str(todate)+'.xlsx'
    # CharterNewHireReport(result_filepath8)
    # print('Processed - Charter New Hire Report')
    
    # result_filepath9 = 'Processed Reports/Charter Extension Report_'+str(todate)+'.xlsx'
    # CharterExtensionReport(result_filepath9)
    # print('Processed - Charter Extension Report')
    
    # result_filepath10 = 'Processed Reports/Charter PERM Report_'+str(todate)+'.xlsx'
    # CharterPermReport(result_filepath10)
    # print('Processed - Charter PERM Report')

    result_filepath11 = 'Processed Reports/Comcast Dashboard Report_'+str(todate)+'.xlsx'
    ComcastDashBoardReport(result_filepath11)
    print('Processed - Comcast DashBoard Report')

    # quit()
    # update dbo.Case 
    # set PrimaryBeneficiaryXref = BeneficiaryXref
    # where BeneficiaryType = 'Primary'
    
    # finalquery

def ActiveEmployeeReport(result_filepath):
    ###################################### Tab 1 Header #############################################
    #Tab 1 - Active Employee Report
    headers = ['Beneficiary Full Name', 'Management Info Employee ID', 'Management Info Department','Management Info Business Unit Code', 'Management Info Dept Group', 'Management Info Job Start Date', 'Birth Country', 'Citizenship', 'Petitioning Job Title', 'Petitioning Job Location', 'Current Status', 'Current Status Expiration Date', 'I-797 Expiration Date', 'I-94 Expiration Date','EAD Expiration', 'AP Expiration', 'Management Info Manager', 'Management Info Second Level Manager', 'NIV Max Out Date', 'Visa Priority Date']

     
    headers_table = ['FullName', 'EmployeeId', 'Department','Business_Unit_Code', 'Department_Group', 'EmploymentStartDate', 'BirthCountry', 'CitizenshipCountry', 'petitioningjobtitle','petitioningjoblocation', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EadExpirationDate','AdvanceParoleExpirationDate', 'ManagerName', 'SecondLevelManager', 'FinalNivDate', 'Priority1Date']
    
    date_columns = ['EmploymentStartDate', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EadExpirationDate','AdvanceParoleExpirationDate', 'FinalNivDate', 'Priority1Date']
    
    # header_names = [{'header': x} for x in headers]
    
    results_active_qry ='''select distinct b.*,
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 
        e.EmployeeId, e.Department, e.Business_Unit_Code, e.Department_Group, e.EmploymentStartDate, e.ManagerName, e.SecondLevelManager,
        c.petitioningjobtitle, c.petitioningjoblocation,bp.Priority1Date 
        FROM dbo.Beneficiary as b 
        LEFT JOIN dbo.[Case] as c on c.BeneficiaryXref=b.BeneficiaryXref
		AND c.SourceCreatedDate = (
		  SELECT
			MAX(SourceCreatedDate)
		  FROM
			dbo.[Case] AS c2
		  WHERE
			c2.BeneficiaryXref = c.BeneficiaryXref
		)
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.[Organization] as o on b.OrganizationXref = o.OrganizationXref
        where b.IsActive = '1'  and o.OrganizationXref = '100000590'
        ORDER BY b.FullName ASC'''
    # organis should charter
    results_active = cursor.execute(results_active_qry).fetchall()

    df = pd.read_sql(results_active_qry,conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    
    for d_h in date_columns:
        if d_h in df:
            if '1900-01-01' in df[d_h]:
                df[d_h] = ''
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter',date_format='MM/DD/YYYY')
    df.to_excel(writer, sheet_name='Active Employee Report', startrow=0, columns=headers, index=False)
    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months='')


     
def ComcastReport(result_filepath):

    # tab_1 Active Beneficiary List

    headers = ['Petitioner', 'Petitioner of Primary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Primary Beneficiary Id', 'Primary Name', 'Relation', 'Birth Country','Citizenship', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'I-129S Expiration ', 'PED Expiration ', 'EAD Type', 'EAD Expiration', 'AP Expiration', 'DS 2019 Expiration ', 'Re-Entry Expiration ', 'Green Card Expiration  ', 'Passport Expiration ', 'Visa Type', 'Visa Expiration ', 'Current Process Type', 'PR Status Method', 'Priority Date', 'Priority Date-Category', 'Priority Date-Country of Chargeability', 'Priority Date-Note', 'Employee ID', 'Job Start Date', 'Job Code', 'Job Title', 'Work Address', 'Job Location City', 'Job Location State', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'PrimaryBeneficiaryXref', 'PrimBenFullName', 'RelationType', 'BirthCountry', 'CitizenshipCountry', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate', 'EADType', 'EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate', 'VisaType', 'VisaExpirationDate', 'CurrentProcessName', 'GreenCardMethod', 'PriorityDate1Date', 'PriorityDate1Category', 'PrioritDate1Country', 'PriorityDate1Note', 'EmployeeId', 'HireDate', 'JobCode', 'JobTitle', 'WorkAddressFull', 'WorkLocationCity', 'WorkLocationState', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number','CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
    

    date_columns = ['CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate','EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate','VisaExpirationDate','PriorityDate1Date','HireDate']


    
    # header_names = [{'header': x} for x in headers]
    

    results_qry ='''select distinct CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,b.I129SEndDate,b.VisaPEDDate,b.EADType,b.EadExpirationDate,b.AdvanceParoleExpirationDate,b.DS2019ExpirationDate,b.REentryPermitExpirationDate,b.GreenCardExpirationDate,b.MostRecentPassportExpirationDate,b.VisaType,b.VisaExpirationDate,c.CurrentProcessName,c.IsCurrentProcess,b.GreenCardMethod,b.PriorityDate1Date,b.PriorityDate1Category,b.PrioritDate1Country,b.PriorityDate1Note,be.EmployeeId,be.HireDate,be.JobCode,be.JobTitle,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName , p2.PetitionerXref
    from [dbo].[Beneficiary] as b
    left join [dbo].[BeneficiaryEmployment] as be on b.BeneficiaryXref = be.BeneficiaryXref
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[Case] as c on c.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
    left join Petitioner p2 on b2.PetitionerXref = p2.PetitionerXref
    where b.IsActive=1 and lower(c.IsCurrentProcess) = 'true' and p2.PetitionerXref = '625365045' order by b2.FullName asc ,  b.BeneficiaryType desc '''
    
    results = cursor.execute(results_qry).fetchall()

    df_tab1 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab1.columns:
        if dfcol not in headers_table:
            df_tab1.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab1 = df_tab1[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab1:
            if '1900-01-01' in df_tab1[d_h]:
                df_tab1[d_h] = ''
            else:
                df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab1.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')
    

    # # ###########################################
    
    # # tab_2 Open Cases - Non-PERM

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name', 'Primary Beneficiary Id', 'Primary Name', 'Relation', 'Process Id', 'Date Opened', 'Process Type', 'Process Reference ', 'Target File Date', 'Case Filed', 'Last Process Activity Completed', 'Primary Process Status','Last Process Activity Date', 'Days Since Last Activity', 'Paralegal', 'Supervisory Paralegal', 'Attorney', 'Summary Case Disposition ', 'Special Instruction Flag', 'HR Special Instruction Flag', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'Employee ID', 'Job Title', 'Petitioning Job Title', 'Petitioning Job Location ', 'Work Address', 'Job Location City', 'Job Location State', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'PrimaryBeneficiaryXref', 'PrimBenFullName', 'RelationType', 'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'targetfiledate', 'CaseFiledDate', 'LastStepCompleted','PrimaryCaseStatus', 'LastStepCompletedDate', 'DaysSinceLastStepCompleted', 'Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'SpecialInstructionFlag', 'SpecialInstructionInfo', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'EmployeeId', 'JobTitle', 'petitioningjobtitle', 'petitioningjoblocation', 'WorkAddressFull', 'WorkLocationCity', 'WorkLocationState', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['SourceCreatedDate','targetfiledate','CaseFiledDate', 'LastStepCompletedDate','CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate']

    # print(date_columns)
    # quit()
    # header_names = [{'header': x} for x in headers]
    

    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName, ' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
    where b.IsActive=1 and c.PrimaryCaseStatus = 'Open' and
    (c.CaseType != 'Labor Cert PERM' and c.CaseType != 'Labor Cert Special Handling' ) and b.PetitionerOfPrimaryBeneficiary = 'Comcast Cable Communications, LLC'
    order by b2.FullName asc ,  b.BeneficiaryType desc'''
        
    results = cursor.execute(results_qry).fetchall()

    df_tab2 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab2.columns:
        if dfcol not in headers_table:
            df_tab2.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab2 = df_tab2[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab2:
            if '1900-01-01' in df_tab2[d_h]:
                df_tab2[d_h] = ''
            else:
                df_tab2[d_h] = pd.to_datetime(df_tab2[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab2.columns = headers #changing dataframe all column names

    ##########################################

    # tab_3 Open Cases - PERM

    #headers as it has to be in O/P file
    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Process Id', 'Date Opened', 'Process Type', 'Process Reference ', 'PERM Memo sent to Employer', 'Prevailing Wage Determination request submitted to DOL', 'Prevailing Wage Determination issued by DOL', 'PWD Expiration Date', 'Recruitment instructions sent to company', 'Target File Date', 'Case Filed', 'Last Process Activity Completed','Primary Process Status', 'Last Process Activity Date', 'Days Since Last Activity', 'Paralegal','SupervisoryParalegal', 'Attorney', 'Summary Case Disposition ', 'Special Instruction Flag', 'HR Special Instruction Flag', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'Employee ID', 'Job Title', 'Petitioning Job Title', 'Petitioning Job Location ', 'Work Address', 'Job Location City', 'Job Location State', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost CtrNo. ','BU Code'] 

    #headers as in dbc
    headers_table =['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'permmemosenttoemployer', 'prevailingwagedeterminationrequestsubmittedtodol', 'prevailingwagedeterminationissuedbydol', 'PwdExpirationDate', 'recruitmentinstructionssenttocompany', 'targetfiledate', 'CaseFiledDate', 'LastStepCompleted', 'PrimaryCaseStatus','LastStepCompletedDate', 'DaysSinceLastStepCompleted', 'Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'SpecialInstructionFlag', 'SpecialInstructionInfo', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'EmployeeId', 'JobTitle', 'petitioningjobtitle', 'petitioningjoblocation', 'WorkAddressFull', 'WorkLocationCity', 'WorkLocationState', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['SourceCreatedDate','PwdExpirationDate','permmemosenttoemployer','recruitmentinstructionssenttocompany','targetfiledate','CaseFiledDate','LastStepCompletedDate','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','FinalNivDate','prevailingwagedeterminationrequestsubmittedtodol','prevailingwagedeterminationissuedbydol']

    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.permmemosenttoemployer,c.prevailingwagedeterminationrequestsubmittedtodol,c.prevailingwagedeterminationissuedbydol,c.PwdExpirationDate,c.recruitmentinstructionssenttocompany,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,c.PrimaryCaseStatus,concat(ParalegalLastName, ' ',ParalegalFirstName) as Paralegal,concat(SupervisoryParalegalLastName,' ', SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(AssociateLastName,' ',AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode  from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    where b.IsActive=1 and c.PrimaryCaseStatus = 'Open' and
    (c.CaseType = 'Labor Cert PERM' or c.CaseType = 'Labor Cert Special Handling' ) and b.PetitionerOfPrimaryBeneficiary = 'Comcast Cable Communications, LLC'
    order by b.FullName asc ,  b.BeneficiaryType desc'''
            
    results = cursor.execute(results_qry).fetchall()

    df_tab3 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab3.columns:
        if dfcol not in headers_table:
            df_tab3.drop(dfcol, axis=1, inplace=True)

    
    # altering the DataFrame - Column order
    df_tab3 = df_tab3[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab3:
            if '1900-01-01' in df_tab3[d_h]:
                df_tab3[d_h] = ''
            else:
                df_tab3[d_h] = pd.to_datetime(df_tab3[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab3.columns = headers #changing dataframe all column names


    ######################################

    


    # tab_4 Document Expiration Report

    #headers as it has to be in O/P file
    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Primary Beneficiary Name', 'Relation', 'Current Status', 'Current Status Expiration Date', 'I-797 Expiration Date ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'I-129S Expiration ', 'PED Expiration ', 'EAD Type', 'EAD Expiration', 'AP Expiration', 'DS 2019 Expiration ', 'Re-Entry Expiration ', 'Green Card Expiration  ', 'Passport Expiration ', 'Visa Type', 'Visa Expiration ', 'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'PrimBenFullName', 'Relationtype', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate', 'EADType', 'EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate', 'VisaType', 'VisaExpirationDate', 'EmployeeId', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']

    date_columns = [ 'CurrentImmigrationStatusExpirationDate2','I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate','EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate','VisaExpirationDate',]
   

  
    results_qry ='''
    select  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.Relationtype ,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,b.I129SEndDate,b.VisaPEDDate,b.EADType,b.EadExpirationDate,b.AdvanceParoleExpirationDate,b.DS2019ExpirationDate,b.REentryPermitExpirationDate,b.GreenCardExpirationDate,b.MostRecentPassportExpirationDate,b.VisaType,b.VisaExpirationDate,be.EmployeeId,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName
    from [dbo].[Beneficiary] as b
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
    where b.IsActive=1 and b.PetitionerOfPrimaryBeneficiary = 'Comcast Cable Communications, LLC'
    order by b.PrimaryBeneficiaryFullName asc ,  b.BeneficiaryType desc'''
            
    results = cursor.execute(results_qry).fetchall()

    df_tab4 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab4.columns:
        if dfcol not in headers_table:
            df_tab4.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab4 = df_tab4[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab4:
            if '1900-01-01' in df_tab4[d_h]:
                df_tab4[d_h] = ''
            else:
                df_tab4[d_h] = pd.to_datetime(df_tab4[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab4.columns = headers #changing dataframe all column names

    exp_month = ((datetime.today())) + relativedelta(days=+240) 
    this_month = ((datetime.today()))
    next_month = ((datetime.today().replace(day=1))) + relativedelta(months=+1) 

    # end_month_str = str(exp_month).split(' ')[0]
    # this_month_str = str(this_month).split(' ')[0]
    # next_month_str = str(next_month).split(' ')[0]


    end_month_str = pd.to_datetime(exp_month)
    this_month_str = pd.to_datetime(this_month)
    next_month_str = pd.to_datetime(next_month).date()


    date_months = ['Current Status Expiration Date','I-797 Expiration Date ','I-94 Expiration Date ','I-129S Expiration ','PED Expiration ','EAD Expiration','AP Expiration','DS 2019 Expiration ','Re-Entry Expiration ','Green Card Expiration  ','Passport Expiration ','Passport Expiration','Visa Expiration ']
  

    df_tab4 = df_tab4[  ((df_tab4['Current Status Expiration Date'] <= end_month_str)&
                         (df_tab4['Current Status Expiration Date'] >= this_month_str))|
                        ((df_tab4['I-797 Expiration Date '] <= end_month_str)&
                         (df_tab4['I-797 Expiration Date '] >= this_month_str))|
                        ((df_tab4['I-94 Expiration Date '] <= end_month_str)&
                         (df_tab4['I-94 Expiration Date '] >= this_month_str))|
                        ((df_tab4['I-129S Expiration '] <= end_month_str)&
                         (df_tab4['I-129S Expiration '] >= this_month_str))|
                        ((df_tab4['PED Expiration '] <= end_month_str)&
                         (df_tab4['PED Expiration '] >= this_month_str))|
                        ((df_tab4['EAD Expiration'] <= end_month_str)&
                         (df_tab4['EAD Expiration'] >= this_month_str))|
                        ((df_tab4['AP Expiration'] <= end_month_str)&
                         (df_tab4['AP Expiration'] >= this_month_str))|
                        ((df_tab4['DS 2019 Expiration '] <= end_month_str)&
                         (df_tab4['DS 2019 Expiration '] >= this_month_str))|
                        ((df_tab4['Re-Entry Expiration '] <= end_month_str)&
                         (df_tab4['Re-Entry Expiration '] >= this_month_str))|
                        ((df_tab4['Green Card Expiration  '] <= end_month_str)&
                         (df_tab4['Green Card Expiration  '] >= this_month_str))|
                        ((df_tab4['Passport Expiration '] <= end_month_str)&
                         (df_tab4['Passport Expiration '] >= this_month_str))|
                        ((df_tab4['Visa Expiration '] <= end_month_str)&
                         (df_tab4['Visa Expiration '] >= this_month_str))]


    # tab_5 Filed and Pending Cases 

    #headers as it has to be in O/P file
    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Target File Date', 'Case Filed', 'Last Process Activity Completed','Primary Process Status', 'Last Process Activity Date', 'Days Since Last Activity', 'Paralegal', 'SupervisoryParalegal', 'Attorney', 'Summary Case Disposition ', 'Special Instruction Flag', 'HR Special Instruction Flag', 'Current Status', 'Current Status Expiration ', 'I-797 Expiration ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'Employee ID', 'Job Title', 'Petitioning Job Title', 'Petitioning Job Location ', 'Work Address', 'Job Location City', 'Job Location State', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost CtrNo. ','BU Code'] 

    #headers as in db
    headers_table =['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription','targetfiledate', 'CaseFiledDate', 'LastStepCompleted', 'PrimaryCaseStatus','LastStepCompletedDate', 'DaysSinceLastStepCompleted', 'Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'SpecialInstructionFlag', 'SpecialInstructionInfo', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'EmployeeId', 'JobTitle', 'petitioningjobtitle', 'petitioningjoblocation', 'WorkAddressFull', 'WorkLocationCity', 'WorkLocationState', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['SourceCreatedDate','PwdExpirationDate','permmemosenttoemployer','recruitmentinstructionssenttocompany','targetfiledate','CaseFiledDate','LastStepCompletedDate','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','FinalNivDate','prevailingwagedeterminationrequestsubmittedtodol','prevailingwagedeterminationissuedbydol']

    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.permmemosenttoemployer,c.prevailingwagedeterminationrequestsubmittedtodol,c.prevailingwagedeterminationissuedbydol,c.PwdExpirationDate,c.recruitmentinstructionssenttocompany,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(ParalegalLastName, ' ',ParalegalFirstName) as Paralegal,concat(SupervisoryParalegalLastName,' ', SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(AssociateLastName,' ',AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    where b.IsActive=1 and c.PrimaryCaseStatus = 'Open' and c.CaseFiledDate != '' and b.PetitionerOfPrimaryBeneficiary = 'Comcast Cable Communications, LLC'
    order by b.FullName asc ,  b.BeneficiaryType desc'''
            
    results = cursor.execute(results_qry).fetchall()

    df_tab5 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab5.columns:
        if dfcol not in headers_table:
            df_tab5.drop(dfcol, axis=1, inplace=True)

    
    # altering the DataFrame - Column order
    df_tab5 = df_tab5[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab5:
            if '1900-01-01' in df_tab5[d_h]:
                df_tab5[d_h] = ''
            else:
                df_tab5[d_h] = pd.to_datetime(df_tab5[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab5.columns = headers #changing dataframe all column names


    ######################################



    # tab_6 PassPort Expiration Report

    #headers as it has to be in O/P file
    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Primary Beneficiary Name', 'Relation', 'Current Status', 'Passport Expiration', 'Employee ID', 'E-Mail ', 'Manager Name', 'Manager E-Mail']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'PrimBenFullName', 'Relationtype', 'Current_Immigration_Status', 'MostRecentPassportExpirationDate', 'EmployeeId', 'WorkEmail', 'ManagerName', 'ManagerEmail']

    date_columns = [ 'CurrentImmigrationStatusExpirationDate2','I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate','EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate','MostRecentPassportExpirationDate','VisaExpirationDate','Current Status Expiration','I-797 Expiration ','I-94 Expiration Date ','I-129S Expiration ','PED Expiration ','EAD Expiration','AP Expiration','DS 2019 Expiration ','Re-Entry Expiration ','Green Card Expiration  ','Passport Expiration ','Visa Expiration ']
   

  
    results_qry ='''
    select  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryFullName,b.Relationtype ,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,b.I129SEndDate,b.VisaPEDDate,b.EADType,b.EadExpirationDate,b.AdvanceParoleExpirationDate,b.DS2019ExpirationDate,b.REentryPermitExpirationDate,b.GreenCardExpirationDate,b.MostRecentPassportExpirationDate,b.VisaType,b.VisaExpirationDate,be.EmployeeId,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b.WorkEmail,be.ManagerEmail,b2.FullName as PrimBenFullName
    from [dbo].[Beneficiary] as b
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
    where b.IsActive=1 and b.PetitionerOfPrimaryBeneficiary = 'Comcast Cable Communications, LLC'
    order by b.PrimaryBeneficiaryFullName asc ,  b.BeneficiaryType desc'''
            
    results = cursor.execute(results_qry).fetchall()

    df_tab6 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab6.columns:
        if dfcol not in headers_table:
            df_tab6.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab6 = df_tab6[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab6:
            if '1900-01-01' in df_tab6[d_h]:
                df_tab6[d_h] = ''
            else:
                df_tab6[d_h] = pd.to_datetime(df_tab6[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab6.columns = headers #changing dataframe all column names

    exp_month = ((datetime.today())) + relativedelta(days=+240) 
    this_month = ((datetime.today()))
    next_month = ((datetime.today().replace(day=1))) + relativedelta(months=+1) 

    # end_month_str = str(exp_month).split(' ')[0]
    # this_month_str = str(this_month).split(' ')[0]
    # next_month_str = str(next_month).split(' ')[0]


    end_month_str = pd.to_datetime(exp_month)
    this_month_str = pd.to_datetime(this_month)
    next_month_str = pd.to_datetime(next_month).date()

    # df_tab6 = df_tab6[((df_tab6['Passport Expiration'] <= end_month_str))]
                    #    (df_tab6['Passport Expiration '] >= this_month_str))]


    df_tab1.to_excel(writer, 'Active Beneficiary List', startrow=0, index=False)
    df_tab2.to_excel(writer, 'Open Cases - Non-PERM', startrow=0, index=False)
    df_tab3.to_excel(writer, 'Open Cases - PERM', startrow=0, index=False)
    df_tab5.to_excel(writer, 'Filed & Pending Cases', startrow=0, index=False)
    df_tab4.to_excel(writer, 'Document Expiration Report', startrow=0, index=False)
    df_tab6.to_excel(writer, 'Passport Expiration Report', startrow=0, index=False)
    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months=date_months)
    

     
def ComcastDashBoardReport(result_filepath):

    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')


    if True: # tab_1 Beneficiary Data

        headers = ['Beneficiary Id', 'Case No. ', 'Beneficiary Type', 'Beneficiary Record Creation Date', 'Beneficiary Status', 'Organization Id', 'Organization Name', 'Petitioner Id', 'Petitioner Name', 'Petitioner of Primary Beneficiary', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Country of Birth', 'Country of Citizenship', 'Immigration Status', 'Employee Id', 'Job Start Date', 'Job Code', 'Current Job Title', 'Work Address-City', 'Work Address-State', 'Work Address-Country', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Center No. ', 'Business Unit Code', 'TPX Project']

        #headers as in db
        headers_table = ['BeneficiaryXref', 'Beneficiary_Xref2', 'BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'OrganizationXref', 'OrganizationName', 'PetitionerXref', 'PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'BirthCountry', 'CitizenshipCountry', 'Current_Immigration_Status', 'EmployeeId', 'HireDate', 'JobCode', 'JobTitle', 'WorkLocationCity', 'WorkLocationState', 'WorkLocationCountry', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode','TPX_PROJECT']
        

        date_columns = ['SourceCreatedDate','CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate','EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate','VisaExpirationDate','PriorityDate1Date','HireDate']


        
        # header_names = [{'header': x} for x in headers]
        

        results_qry ='''select distinct CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,b.BeneficiaryXref,b.Beneficiary_Xref2,b.BeneficiaryType,b.SourceCreatedDate,b.IsActive,o.OrganizationXref,o.OrganizationName,p.PetitionerXref,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.LastName,b.FirstName,b.PrimaryBeneficiaryXref,b2.LastName as PrimaryBeneficiaryLastName ,b2.FirstName as PrimaryBeneficiaryFirstName,b.RelationType,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,be.EmployeeId,be.HireDate,be.JobCode,be.JobTitle,be.WorkLocationCity,be.WorkLocationState,be.WorkLocationCountry,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,'' as TPX_PROJECT,b2.FullName


        from [dbo].[Beneficiary] as b
        left join [dbo].[BeneficiaryEmployment] as be on b.BeneficiaryXref = be.BeneficiaryXref
        left join  [dbo].[Organization] as o on b.OrganizationXref = o.OrganizationXref 
        left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
        left join [dbo].[Case] as c on c.BeneficiaryXref = b.BeneficiaryXref
        left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
        left join Petitioner p2 on b2.PetitionerXref = p2.PetitionerXref

        where p2.PetitionerXref = '625365045' and 
            
            (b.IsActive=1 or   
            (b.BeneficiaryType = 'primary' and (DATEDIFF(year,be.HireDate,getdate()) <=2)))

        order by b2.FullName asc ,  b.BeneficiaryType desc '''

        # lower(c.IsCurrentProcess) = 'true' and 
        # p2.PetitionerXref = '625365045' 
        
        results = cursor.execute(results_qry).fetchall()

        df_tab1 = pd.read_sql(results_qry,conn)

        for dfcol in df_tab1.columns:
            if dfcol not in headers_table:
                df_tab1.drop(dfcol, axis=1, inplace=True)
        
        # altering the DataFrame - Column order
        df_tab1 = df_tab1[headers_table]
        
        for d_h in date_columns:
            if d_h in df_tab1:
                if '1900-01-01' in df_tab1[d_h]:
                    df_tab1[d_h] = ''
                else:
                    df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    
        df_tab1.columns = headers #changing dataframe all column names
    
    # # ###########################################
    
    if True: # tab_2 Case Data

        headers = ['Beneficiary Id', 'Case No.', 'Beneficiary Type', 'Beneficiary Record Creation Date', 'Beneficiary Status', 'Organization Id', 'Organization Name', 'Petitioner Id', 'Petitioner Name', 'Petitionerof Primary Beneficiary', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Immigration Status', 'Immigration Status Exp Date', 'Process Opened Date', 'Process Id', 'Process Type', 'Process Reference', 'Process Filed Date', 'Primary Process Status', 'Secondary Process Status', 'Secondary Process Status Date', 'Final Action', 'Final Action Date', 'Case Closed', 'Employee Id', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Center No. ', 'Business Unit Code', 'TPX Project']

        

        #headers as in db
        headers_table = ['BeneficiaryXref', 'Beneficiary_Xref2', 'BeneficiaryType', 'BenSourceCreatedDate', 'BeneficiaryRecordStatus', 'OrganizationXref', 'OrganizationName', 'PetitionerXref', 'PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'CaseSourceCreatedDate', 'CaseXref', 'CaseType', 'CaseDescription', 'CaseFiledDate', 'PrimaryCaseStatus', 'LastStepCompleted', 'LastStepCompletedDate', 'FinalAction', 'FinalActionDate', 'CaseClosedDate', 'EmployeeId', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode','TPX_PROJECT']
            

        date_columns = ['CaseClosedDate','CaseSourceCreatedDate', 'BenSourceCreatedDate','targetfiledate','CaseFiledDate', 'LastStepCompletedDate','CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate','FinalActionDate']

        # print(date_columns)
        # quit()
        # header_names = [{'header': x} for x in headers]
        

        results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,
        
        b.BeneficiaryXref,b.Beneficiary_Xref2,b.BeneficiaryType,b.SourceCreatedDate as BenSourceCreatedDate,b.IsActive,o.OrganizationXref,o.OrganizationName,p.PetitionerXref,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.LastName,b.FirstName,b.PrimaryBeneficiaryXref,b2.PrimaryBeneficiaryLastName,b2.PrimaryBeneficiaryFirstName,b.RelationType,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,c.SourceCreatedDate as CaseSourceCreatedDate,c.CaseXref,c.CaseType,c.CaseDescription,c.CaseFiledDate,c.PrimaryCaseStatus,c.LastStepCompleted,c.LastStepCompletedDate,c.FinalAction,c.FinalActionDate,c.CaseClosedDate,be.EmployeeId,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,'' as TPX_PROJECT,b2.FullName

        from [dbo].[Case] as c
        left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
        left join [dbo].[Organization] as o on b.OrganizationXref = o.OrganizationXref 
        left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
        left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
        left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
        left join Petitioner p2 on b2.PetitionerXref = p2.PetitionerXref

        where p2.PetitionerXref = '625365045' and
                
                ((DATEDIFF(YEAR,c.SourceCreatedDate,GETDATE())<=2) or
                (DATEDIFF(YEAR,c.CaseFiledDate,GETDATE())<=2) or
                (c.PrimaryCaseStatus = 'open' and b.IsActive=1) or
                
				(c.FinalAction in ('granted','denied') and
                (DATEDIFF(YEAR,c.FinalActionDate,GETDATE())<=1))) 

        order by b2.FullName asc ,  b.BeneficiaryType desc'''
    
        # where   ((DATEDIFF(YEAR,c.CaseOpenDate,GETDATE())<=3) or
        # 		(DATEDIFF(YEAR,c.CaseFiledDate,GETDATE())<=3) or
        # 		(c.PrimaryCaseStatus = 'open' and b.IsActive=1)) or
                
        # 		(c.FinalAction in ('granted','denied') or
        # 		(DATEDIFF(YEAR,c.FinalActionDate,GETDATE())<=2))
            
        # results = cursor.execute(results_qry).fetchall()

        df_tab2 = pd.read_sql(results_qry,conn)

        df_tab2 = df_tab2[headers_table]

        for dfcol in df_tab2.columns:
            if dfcol not in headers_table:
                df_tab2.drop(dfcol, axis=1, inplace=True)
        
        # altering the DataFrame - Column order
        df_tab2 = df_tab2[headers_table]

        for d_h in date_columns:
            if d_h in df_tab2:
                if '1900-01-01' in df_tab2[d_h]:
                    df_tab2[d_h] = ''
                else:
                    df_tab2[d_h] = pd.to_datetime(df_tab2[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    
        df_tab2.columns = headers #changing dataframe all column names

    ##########################################

    if True: # tab_3 SLA - Case Milestones

        headers = ['Beneficiary Id', 'Case No.', 'Beneficiary Type', 'Beneficiary Record Creation Date', 'Beneficiary Status', 'Organization', 'Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Immigration Status', 'Immigration Status Exp Date', 'Process Opened Date','Case Petition Id', 'Process Id', 'Process Type', 'Process Reference', 'Process Filed Date', 'Primary Process Status', 'Secondary Process Status', 'Secondary Process Status Date', 'Final Action', 'Final Action Date', 'Case Closed', 'Employee Id', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'Business Unit Code', 'VP', 'TPX Project', 'Questionnaire Sent to Manager', 'Questionnaire Sent to FN', 'Questionnaire Returned by Manager', 'Questionnaire Returned by FN', 'All Petitioning Company Info Received', 'All FN Docs Received', 'LCA Filed', 'Documents Sent for Signature', 'Singed Docs Received', 'AOS Docs Sent for Signature', 'Signed AOS Docs Received', 'RFE Received', 'RFE Due Date', 'RFE Docs Requested', 'RFE Docs Received', 'RFE Docs to ER for Review / Signature', 'RFE Response Submitted', 'PERM Memo Sent to Employer', 'Approval of PERM Memo Received', 'Employee Work Experience Chart Sent', 'Employee Work Experience Chart Received', 'PWD Request Submitted to DOL', 'PWD Issued by DOL', 'PWD Expiration Date', 'Recruitment Approval Received from ER', 'Recruitment Instructions Sent to Company', 'Job Order Placed with SWA', 'Dated Copies of All Recruitment Received', 'Recruitment Report Sent to Company', 'Recruitment Report Received', 'Form 9089 Sent to FN and Employer', 'Edits to Form 9089 Received from FN and Employer', 'Form 9089 Submitted to DOL', 'Audit Notice Received', 'Audit Docs to ER for Review / Signature', 'Audit Docs Received from ER', 'Audit Response Sent to DOL']

        

        #headers as in db
        headers_table = ['BeneficiaryXref', 'Beneficiary_Xref2', 'BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'OrganizationName', 'PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'ProcessCreatedDate','CasePetitionId', 'CaseXref', 'CaseType', 'CaseDescription', 'CaseFiledDate', 'PrimaryCaseStatus', 'LastStepCompleted', 'LastStepCompletedDate', 'FinalAction', 'FinalActionDate', 'CaseClosedDate', 'EmployeeId', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode', 'SecondLevelManager', 'TPX_PROJECT', 'questionnairesenttomanager', 'questionnairessenttofn', 'questionnairecompletedandreturnedbymanager', 'questionnairecompletedandreturnedbyfn', 'allpetitioningcompanyinforeceived', 'allfndocsreceived', 'lcafiled', 'formsanddocumentationsubmittedforsignature', 'signedformsandletterreceived', 'dateaosformssentforsignature', 'datesignedaosformsreceived', 'RFEAuditReceivedDate', 'RFEAuditDueDate', 'RFEDocsReqestedDate', 'RFEDocsReceivedDate', 'RFE_Docs_to_ER_for_Review_Signature', 'RFEAuditSubmittedDate', 'permmemosenttoemployer', 'approvalofpermmemoreceived', 'employeeworkexperiencechartsent', 'employeeworkexperiencechartreceived', 'prevailingwagedeterminationrequestsubmittedtodol', 'prevailingwagedeterminationissuedbydol', 'PwdExpirationDate', 'Recruitment_Approval_Received_from_ER', 'recruitmentinstructionssenttocompany', 'joborderplacedwithswa', 'datedcopiesofallrecruitmentreceived', 'recruitmentreportsenttocompany', 'recruitmentreportreceived', 'form9089senttofnandemployer', 'editstoform9089receivedfromfnandemployer', 'form9089submittedtodol', 'PERMAuditReceivedDate', 'Audit_Docs_to_ER_for_Review_Signature', 'Audit_Docs_Received_from_ER', 'PERMAuditSubmittedDate']
        

        date_columns = ['SourceCreatedDate', 'ProcessCreatedDate','CurrentImmigrationStatusExpirationDate2', 'SourceCreatedDate', 'CaseFiledDate', 'LastStepCompletedDate', 'FinalActionDate', 'CaseClosedDate', 'questionnairessenttofn', 'questionnairesenttomanager', 'questionnairecompletedandreturnedbymanager', 'questionnairecompletedandreturnedbyfn', 'allpetitioningcompanyinforeceived', 'allfndocsreceived', 'lcafiled', 'formsanddocumentationsubmittedforsignature', 'signedformsandletterreceived', 'dateaosformssentforsignature', 'datesignedaosformsreceived', 'RFEAuditReceivedDate','RFE_Docs_to_ER', 'RFEAuditDueDate', 'RFEDocsReqestedDate', 'RFEDocsReceivedDate', 'RFEAuditSubmittedDate', 'permmemosenttoemployer', 'approvalofpermmemoreceived', 'employeeworkexperiencechartsent', 'employeeworkexperiencechartreceived', 'prevailingwagedeterminationrequestsubmittedtodol', 'prevailingwagedeterminationissuedbydol', 'PwdExpirationDate', 'recruitmentinstructionssenttocompany', 'joborderplacedwithswa', 'datedcopiesofallrecruitmentreceived', 'recruitmentreportsenttocompany', 'recruitmentreportreceived', 'form9089senttofnandemployer', 'editstoform9089receivedfromfnandemployer', 'form9089submittedtodol', 'PERMAuditReceivedDate', 'PERMAuditSubmittedDate']

        
        # header_names = [{'header': x} for x in headers]
        

        results_qry ='''select distinct CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,b.BeneficiaryXref,b.Beneficiary_Xref2,b.BeneficiaryType,b.SourceCreatedDate,b.IsActive,o.OrganizationName,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.LastName,b.FirstName,b.PrimaryBeneficiaryXref,b2.LastName as PrimaryBeneficiaryLastName,b2.FirstName as PrimaryBeneficiaryFirstName,b.RelationType,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,c.SourceCreatedDate as ProcessCreatedDate,c.CasePetitionId,c.CaseXref,c.CaseType,c.CaseDescription,c.CaseFiledDate,c.PrimaryCaseStatus,c.LastStepCompleted,c.LastStepCompletedDate,c.FinalAction,c.FinalActionDate,c.CaseClosedDate,be.EmployeeId,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,be.SecondLevelManager,'' as TPX_PROJECT,c.questionnairesenttomanager,c.questionnairessenttofn,c.questionnairecompletedandreturnedbymanager,c.questionnairecompletedandreturnedbyfn,c.allpetitioningcompanyinforeceived,c.allfndocsreceived,c.lcafiled,c.formsanddocumentationsubmittedforsignature,c.signedformsandletterreceived,c.dateaosformssentforsignature,c.datesignedaosformsreceived,c.RFEAuditReceivedDate,c.RFEAuditDueDate,c.RFEDocsReqestedDate,c.RFEDocsReceivedDate,'' as RFE_Docs_to_ER_for_Review_Signature,c.RFEAuditSubmittedDate,c.permmemosenttoemployer,c.approvalofpermmemoreceived,c.employeeworkexperiencechartsent,c.employeeworkexperiencechartreceived,c.prevailingwagedeterminationrequestsubmittedtodol,c.prevailingwagedeterminationissuedbydol,c.PwdExpirationDate,'' as Recruitment_Approval_Received_from_ER,c.recruitmentinstructionssenttocompany,c.joborderplacedwithswa,c.datedcopiesofallrecruitmentreceived,c.recruitmentreportsenttocompany,c.recruitmentreportreceived,c.form9089senttofnandemployer,c.editstoform9089receivedfromfnandemployer,c.form9089submittedtodol,c.PERMAuditReceivedDate,'' as Audit_Docs_to_ER_for_Review_Signature,'' as Audit_Docs_Received_from_ER,c.PERMAuditSubmittedDate,b2.FullName

        from [dbo].[Case] as c
        left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
        left join [dbo].[Organization] as o on b.OrganizationXref = o.OrganizationXref 
        left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
        left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
        left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
        left join Petitioner p2 on b2.PetitionerXref = p2.PetitionerXref


		where p2.PetitionerXref = '625365045' and
        
          ((c.PrimaryCaseStatus = 'open' and c.CasePetitionId in ('100003008','100003034','100003010','100003009','100003013')) or
		
	      (c.CasePetitionId = '100003008' and c.PrimaryCaseStatus = 'closed' and  (datediff(year,c.SourceCreatedDate,getdate()) <1)) or

		  (c.CasePetitionId = '100003034' and
		   c.CaseDescription in ('Change of Employer','COE', 'H-1B Change COE', 'H-1B Change of Employer', 'H-1B Change of ER' ,'Ext', 'Extension', 'H-1B Ext', 'H-1B Extension') and
		   c.PrimaryCaseStatus = 'closed' and
		   (datediff(year,c.SourceCreatedDate,getdate()) <=2)) or

		   (c.CasePetitionId = '100003010' and
		   c.PrimaryCaseStatus = 'closed' and
		   (datediff(year,c.SourceCreatedDate,getdate()) <=2)) or

		   (c.CasePetitionId = '100003009' and
		   c.PrimaryCaseStatus = 'closed' and
		   (datediff(year,c.SourceCreatedDate,getdate()) <=2)) or	

		   (c.CasePetitionId = '100003013' and
		   c.PrimaryCaseStatus = 'closed' and
		   (datediff(year,c.SourceCreatedDate,getdate()) <=2)) or	

		   (c.CaseType != 'Labor Cert PERM' and
		   (datediff(year,c.RFEAuditReceivedDate,getdate()) <=2)) or	

		   (c.CaseType = 'Labor Cert PERM' and
		   (datediff(year,c.PERMAuditReceivedDate,getdate()) <=2)))	


        order by b2.FullName asc ,  b.BeneficiaryType desc'''

    
        
        # results = cursor.execute(results_qry).fetchall()

        df_tab3 = pd.read_sql(results_qry,conn)
        for dfcol in df_tab3.columns:
            if dfcol not in headers_table:
                df_tab3.drop(dfcol, axis=1, inplace=True)
        
        # altering the DataFrame - Column order
        df_tab3 = df_tab3[headers_table]
        
        for d_h in date_columns:
            if d_h in df_tab3:
                if '1900-01-01' in df_tab3[d_h]:
                    df_tab3[d_h] = ''
                else:
                    df_tab3[d_h] = pd.to_datetime(df_tab3[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    
        df_tab3.columns = headers #changing dataframe all column names
        writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')

    df_tab1.to_excel(writer, 'Beneficiary Data', startrow=0, index=False)
    df_tab2.to_excel(writer, 'Case Data', startrow=0, index=False)
    df_tab3.to_excel(writer, 'SLA - Case Milestones', startrow=0, index=False)

    writer.save()
    writer.close()
    
    # add_designs(file_path=result_filepath,date_months=date_months)
    
  
def CharterActiveBenReport(result_filepath):

    # tab_1 CharterActiveBeneficiaryReport

    headers = ['Petitioner', 'Petitioner of Primary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Primary Beneficiary Id', 'Primary Name', 'Relation', 'Birth Country','Citizenship', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ', 'NIV Max Out Date', 'I-129S Expiration ', 'PED Expiration ', 'EAD Type', 'EAD Expiration', 'AP Expiration','Current Process Type', 'PR Status Method', 'Priority Date', 'Priority Date-Category', 'Priority Date-Country of Chargeability', 'Priority Date-Note', 'Employee ID', 'Job Start Date', 'Job Code', 'Job Title', 'Work Address', 'Job Location City', 'Job Location State', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code','VP']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName', 'PrimaryBeneficiaryXref', 'PrimBenFullName', 'RelationType', 'BirthCountry', 'CitizenshipCountry', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate', 'EADType', 'EadExpirationDate', 'AdvanceParoleExpirationDate','CurrentProcessName', 'GreenCardMethod', 'PriorityDate1Date', 'PriorityDate1Category', 'PrioritDate1Country', 'PriorityDate1Note', 'EmployeeId', 'HireDate', 'JobCode', 'JobTitle', 'WorkAddressFull', 'WorkLocationCity', 'WorkLocationState', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number','CostCenter', 'CostCenterNumber', 'BusinessUnitCode','SecondLevelManager']
    

    date_columns = ['CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'FinalNivDate', 'I129SEndDate', 'VisaPEDDate','EadExpirationDate', 'AdvanceParoleExpirationDate', 'DS2019ExpirationDate', 'REentryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate','VisaExpirationDate','PriorityDate1Date','HireDate']


    
    # header_names = [{'header': x} for x in headers]
    

    results_qry ='''select distinct CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,b.OrganizationXref,p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,b.I129SEndDate,b.VisaPEDDate,b.EADType,b.EadExpirationDate,b.AdvanceParoleExpirationDate,b.DS2019ExpirationDate,b.REentryPermitExpirationDate,b.GreenCardExpirationDate,b.MostRecentPassportExpirationDate,b.VisaType,b.VisaExpirationDate,c.CurrentProcessName,c.IsCurrentProcess,b.GreenCardMethod,b.PriorityDate1Date,b.PriorityDate1Category,b.PrioritDate1Country,b.PriorityDate1Note,be.EmployeeId,be.HireDate,be.JobCode,be.JobTitle,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.SecondLevelManager,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName , p2.PetitionerXref
    from [dbo].[Beneficiary] as b
    left join [dbo].[BeneficiaryEmployment] as be on b.BeneficiaryXref = be.BeneficiaryXref
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[Case] as c on c.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref
    left join Petitioner p2 on b2.PetitionerXref = p2.PetitionerXref
    
    where b.IsActive=1 and 
	lower(b.BeneficiaryType) = 'primary' and
	lower(b.Current_Immigration_Status) not in ('cpr', 'lpr','permanent residence') and
	

    b.OrganizationXref = '100000590' and 
    lower(b.PetitionerofPrimaryBeneficiary) = 'charter communications, inc.'
    order by PrimBenFullName asc ,  b.BeneficiaryType desc
    '''
    # add below cond if required
    # lower(c.IsCurrentProcess) = 'true' and 
    
    results = cursor.execute(results_qry).fetchall()

    df_tab1 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab1.columns:
        if dfcol not in headers_table:
            df_tab1.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab1 = df_tab1[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab1:
            if '1900-01-01' in df_tab1[d_h]:
                df_tab1[d_h] = ''
            else:
                df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab1.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')
    
    # Condition 4: Where “Current Status = TN-1; TN-2; TN Visa; TN-Canada, TN-Mexico ” – replace value with “TN”
    for index, value in enumerate(df_tab1['Current Status']):
        if str(value).strip().lower() in ['tn-1','tn-2','tn visa','tn-canada','tn-mexico','tn-visa']:
            df_tab1['Current Status'][index] = 'TN'



    df_tab1.to_excel(writer, 'Active Beneficiary List', startrow=0, index=False)
    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months='')

def CharterNewHireReport(result_filepath):

    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')

    # # ###########################################
    
    # # tab_1 H-1B COE & Amendment Report

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name',  'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Online Intake Date','All FN Docs Received','All Petitioning Company Info Received', 'Target File Date', 'Case Filed','Receipt Date',	'RFE Received',	'RFE Response Submitted', 'Date Status Valid From', 'Date Status Valid To',	'Final Action',	'Approval Package Sent', 'Last Process Activity Completed','Last Process Activity Date','Paralegal', 'Supervisory Paralegal', 'Attorney', 'Summary Case Disposition ', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ',  'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName',  'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'OnlineIntakeDate','allfndocsreceived','allpetitioningcompanyinforeceived','targetfiledate', 'CaseFiledDate','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditSubmittedDate','CaseValidFromDate','CaseExpirationDate','FinalAction','approvalpackagesent','LastStepCompleted', 'LastStepCompletedDate','Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EmployeeId',  'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['approvalpackagesent','allpetitioningcompanyinforeceived','allfndocsreceived','CaseReceivedDate', 'RFEAuditSubmittedDate', 'ImmigrationStatusExpirationDate', 'CurrentImmigrationStatusExpirationDate2', 'FinalNivDate', 'CaseValidFromDate', 'CaseFiledDate', 'I797ExpirationDate', 'CaseExpirationDate', 'LastStepCompletedDate', 'SourceCreatedDate', 'OnlineIntakeDate', 'RFEAuditReceivedDate', 'targetfiledate','']

    # print(date_columns)
    # quit()
    # header_names = [{'header': x} for x in headers]
    

    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,c.OnlineIntakeDate,c.allfndocsreceived,c.allpetitioningcompanyinforeceived,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName,' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditSubmittedDate,c.CaseValidFromDate,c.CaseExpirationDate,c.FinalAction,c.approvalpackagesent
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref

    where b.IsActive=1 and 
    c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId = '100003034' and

    lower(c.CaseDescription) in ('change of employer','amend','amendment (with extension)','amendment (without extension)','h-1b','h-1b amend','h-1b change of er','please select','*please select','null','') and

    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590'
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''

    
        
    results = cursor.execute(results_qry).fetchall()

    df_tab1 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab1.columns:
        if dfcol not in headers_table:
            df_tab1.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab1 = df_tab1[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab1:
            if '1900-01-01' in df_tab1[d_h]:
                df_tab1[d_h] = ''
            else:
                df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab1.columns = headers #changing dataframe all column names


    # # ###########################################
    
    # # tab_2 TN Reports

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name',  'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Online Intake Date','All FN Docs Received','All Petitioning Company Info Received', 'Target File Date', 'Case Filed','Receipt Date','Client Prepped for Interview','Date of POE Consular Interview','Date Status Valid From', 'Date Status Valid To','Final Action','Received Copy of I-94 and Visa Stamp','Last Process Activity Completed','Last Process Activity Date','Paralegal', 'Supervisory Paralegal', 'Attorney', 'Summary Case Disposition ', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ',  'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName',  'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'OnlineIntakeDate','allfndocsreceived','allpetitioningcompanyinforeceived','targetfiledate', 'CaseFiledDate','CaseReceivedDate','BeneficiaryPreppedForInterview','consularinterviewdate','CaseValidFromDate','CaseExpirationDate','FinalAction','ReceivedCopyOf_I94_VisaStamp','LastStepCompleted', 'LastStepCompletedDate','Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EmployeeId',  'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['approvalpackagesent','allpetitioningcompanyinforeceived','allfndocsreceived','CaseFiledDate', 'CaseExpirationDate', 'CaseReceivedDate', 'OnlineIntakeDate', 'CurrentImmigrationStatusExpirationDate2', 'targetfiledate', 'LastStepCompletedDate', 'CaseValidFromDate', 'FinalNivDate', 'consularinterviewdate', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'SourceCreatedDate']


    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,c.OnlineIntakeDate,c.allfndocsreceived,c.allpetitioningcompanyinforeceived,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName, ' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditSubmittedDate,c.CaseValidFromDate,c.CaseExpirationDate,c.FinalAction,c.approvalpackagesent,c.BeneficiaryPreppedForInterview,c.consularinterviewdate,c.ReceivedCopyOf_I94_VisaStamp
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref


    where b.IsActive=1 and c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId = '100003043' and

    lower(c.CaseDescription) in ('amendment (with extension)','change of employer','initial','please select','*Please select','tn border','tn/canada','tn/mexico', 'Null') and

    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590' 
    
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''
        
    results = cursor.execute(results_qry).fetchall()

    df_tab2 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab2.columns:
        if dfcol not in headers_table:
            df_tab2.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab2 = df_tab2[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab2:
            if '1900-01-01' in df_tab2[d_h]:
                df_tab2[d_h] = ''
            else:
                df_tab2[d_h] = pd.to_datetime(df_tab2[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab2.columns = headers #changing dataframe all column names


    # # ###########################################
    



    # # tab_3 Pre-Hire Assessments
    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name',  'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Summary Case Disposition ','Pre-onboarding questionnaire received','Analysis of pre-onboarding questionnaire sent','Prevailing Wage assessment sent to recruiter', 'Current Status', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code','Recruiter']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName',  'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'CaseComments','PreOnboardingQuestionnaireReceived','AnalysisOfPreOnboardingQuestionnaireSent','prevailingwagedeterminationrequestsubmittedtodol', 'Current_Immigration_Status', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode','ClientBillingCode']
        

    date_columns = ['approvalpackagesent','allpetitioningcompanyinforeceived','allfndocsreceived','SourceCreatedDate']


    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,c.OnlineIntakeDate,c.allfndocsreceived,c.allpetitioningcompanyinforeceived,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName, ' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName,c.ReceiptDateReceivedByGovt,c.RFEAuditReceivedDate,c.RFEAuditSubmittedDate,c.CaseValidFromDate,c.CaseExpirationDate,c.FinalAction,c.approvalpackagesent,c.BeneficiaryPreppedForInterview,c.consularinterviewdate,c.ReceivedCopyOf_I94_VisaStamp,c.ClientBillingCode,c.PreOnboardingQuestionnaireReceived,c.AnalysisOfPreOnboardingQuestionnaireSent,c.prevailingwagedeterminationrequestsubmittedtodol
    
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref

    where b.IsActive=1 and c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId = '610010137' and
    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590'
    
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''
        
    results = cursor.execute(results_qry).fetchall()

    df_tab3 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab3.columns:
        if dfcol not in headers_table:
            df_tab3.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab3 = df_tab3[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab3:
            if '1900-01-01' in df_tab3[d_h]:
                df_tab3[d_h] = ''
            else:
                df_tab3[d_h] = pd.to_datetime(df_tab3[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab3.columns = headers #changing dataframe all column names

    ##########################################

    df_tab1.to_excel(writer, 'H-1B COE & Amendment Report', startrow=0, index=False)
    df_tab2.to_excel(writer, 'TN Report', startrow=0, index=False)
    df_tab3.to_excel(writer, 'Pre-Hire Assessments', startrow=0, index=False)

    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months='')

def CharterExtensionReport(result_filepath):

    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')

    # # ###########################################
    
    # # tab_1 H-1B Extension Report

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name',  'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Online Intake Date','All FN Docs Received','All Petitioning Company Info Received', 'Target File Date', 'Case Filed','Receipt Date',	'RFE Received',	'RFE Response Submitted', 	'Final Action',	'Last Process Activity Completed','Last Process Activity Date','Paralegal', 'Supervisory Paralegal', 'Attorney', 'Summary Case Disposition ', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ',  'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName',  'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'OnlineIntakeDate','allfndocsreceived','allpetitioningcompanyinforeceived','targetfiledate', 'CaseFiledDate','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditSubmittedDate','FinalAction','LastStepCompleted', 'LastStepCompletedDate','Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EmployeeId',  'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['approvalpackagesent','allpetitioningcompanyinforeceived','allfndocsreceived','CaseReceivedDate', 'RFEAuditSubmittedDate', 'ImmigrationStatusExpirationDate', 'CurrentImmigrationStatusExpirationDate2', 'FinalNivDate', 'CaseValidFromDate', 'CaseFiledDate', 'I797ExpirationDate', 'CaseExpirationDate', 'LastStepCompletedDate', 'SourceCreatedDate', 'OnlineIntakeDate', 'RFEAuditReceivedDate', 'targetfiledate','']

    # print(date_columns)
    # quit()
    # header_names = [{'header': x} for x in headers]
    

    # # tab_1 H-1B Extension Report
    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,c.OnlineIntakeDate,c.allfndocsreceived,c.allpetitioningcompanyinforeceived,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName,' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditSubmittedDate,c.CaseValidFromDate,c.CaseExpirationDate,c.FinalAction,c.approvalpackagesent
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref

    where b.IsActive=1 and 
    c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId = '100003034' and

    lower(c.CaseDescription) in ('extension') and

    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590'
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''

    results = cursor.execute(results_qry).fetchall()

    df_tab1 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab1.columns:
        if dfcol not in headers_table:
            df_tab1.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab1 = df_tab1[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab1:
            if '1900-01-01' in df_tab1[d_h]:
                df_tab1[d_h] = ''
            else:
                df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab1.columns = headers #changing dataframe all column names


    # # ###########################################
    
    # # tab_2 TN Extension Report

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No. ', 'Beneficiary Name',  'Process Id', 'Date Opened', 'Process Type', 'Process Reference ','Online Intake Date','All FN Docs Received','All Petitioning Company Info Received', 'Target File Date', 'Case Filed','Receipt Date',	'RFE Received',	'RFE Response Submitted', 	'Final Action',	'Last Process Activity Completed','Last Process Activity Date','Paralegal', 'Supervisory Paralegal', 'Attorney', 'Summary Case Disposition ', 'Current Status', 'Current Status Expiration', 'I-797 Expiration Date ', 'I-94 Expiration Date ',  'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No. ', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2', 'FullName',  'CaseXref', 'SourceCreatedDate', 'CaseType', 'CaseDescription', 'OnlineIntakeDate','allfndocsreceived','allpetitioningcompanyinforeceived','targetfiledate', 'CaseFiledDate','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditSubmittedDate','FinalAction','LastStepCompleted', 'LastStepCompletedDate','Paralegal', 'SupervisoryParalegal', 'Attorney', 'CaseComments', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'I797ExpirationDate', 'ImmigrationStatusExpirationDate', 'EmployeeId',  'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['approvalpackagesent','allpetitioningcompanyinforeceived','allfndocsreceived','CaseReceivedDate', 'RFEAuditSubmittedDate', 'ImmigrationStatusExpirationDate', 'CurrentImmigrationStatusExpirationDate2', 'FinalNivDate', 'CaseValidFromDate', 'CaseFiledDate', 'I797ExpirationDate', 'CaseExpirationDate', 'LastStepCompletedDate', 'SourceCreatedDate', 'OnlineIntakeDate', 'RFEAuditReceivedDate', 'targetfiledate','']

    # print(date_columns)
    # quit()
    # header_names = [{'header': x} for x in headers]
    
    # # tab_2 TN Extension Report
    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus,p.PetitionerName,c.OnlineIntakeDate,c.allfndocsreceived,c.allpetitioningcompanyinforeceived,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,b.PrimaryBeneficiaryXref,b.RelationType,c.CaseXref,c.SourceCreatedDate,c.CaseType,c.CaseDescription,c.targetfiledate,c.CaseFiledDate,c.LastStepCompleted,c.LastStepCompletedDate,c.DaysSinceLastStepCompleted,concat(c.ParalegalLastName,' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,c.CaseComments,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.PrimaryCaseStatus,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.FinalNivDate,be.EmployeeId,be.JobTitle,c.petitioningjobtitle,c.petitioningjoblocation,be.WorkAddressFull,be.WorkLocationCity,be.WorkLocationState,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditSubmittedDate,c.CaseValidFromDate,c.CaseExpirationDate,c.FinalAction,c.approvalpackagesent
    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref

    where b.IsActive=1 and 
    c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId in ('100003034','100003044') and

    lower(c.CaseDescription) in ('extension') and

    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590'
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''

    results = cursor.execute(results_qry).fetchall()

    df_tab2 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab2.columns:
        if dfcol not in headers_table:
            df_tab2.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab2 = df_tab2[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab2:
            if '1900-01-01' in df_tab2[d_h]:
                df_tab2[d_h] = ''
            else:
                df_tab2[d_h] = pd.to_datetime(df_tab2[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab2.columns = headers #changing dataframe all column names


    # # ###########################################

    df_tab1.to_excel(writer, 'H-1B Extension Report', startrow=0, index=False)
    df_tab2.to_excel(writer, 'TN Extension Report', startrow=0, index=False)

    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months='')



def CharterPermReport(result_filepath):

    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='MM/DD/YYYY')

    # # ###########################################
    
    # # tab_1 CharterPermReport

    headers = ['Petitioner', 'Petitioner of Primary Beneficiary', 'Beneficiary Id (LLX)', 'Beneficiary Type', 'Case No.', 'Beneficiary Name', 'Process  Opened Date', 'Process Id', 'Process Type', 'Process Reference', 'Target File Date', 'Summary Case Disposition', 'Last Process Activity', 'Last Process Activity Date', 'Special Instruction Flag', 'HR Special Instructions Flag', 'Days Since Last Activity', 'Petitioning Job Title', 'Petitioning Job Location', 'Online Intake Date', 'Questionnaire sent to manager', 'Questionnaires Sent to FN', 'Questionnaire completed and returned by Manager', 'Questionnaire completed and returned by FN', 'PERM Memo sent to employer', 'Approval of PERM Memo received', 'Employee Work Experience Chart sent', 'Employee Work Experience Chart received', 'Employment Verification Letters sent to employee', 'Signed Employment Verification Letters received', 'Prevailing Wage Determination request submitted to DOL', 'Prevailing Wage Determination issued by DOL', 'PWD expiration date', 'Recruitment instructions sent to company','1st Additional Recruitment Step Placed', 'Dated copies of all recruitment received', 'Completed evaluation questionnaires and resumes received', 'Recruitment report sent to company', 'Recruitment Report Received', 'Form 9089 sent to FN and Employer', 'Edits to Form 9089 received from FN and Employer', 'Form 9089 submitted to DOL', 'Paralegal Name', 'Supervisory Paralegal', 'Attorney', 'Birth Country', 'Citizenship Country', 'Current Status', 'Current Status Expires', 'I94 Expires', 'I-797 Expires', 'NIV Max Out Date', 'Visa Priority Date', 'Employee ID', 'Manager Name', 'Business Partner Name', 'Dept', 'Dept Group', 'Dept Number', 'Cost Center', 'Cost Ctr No.', 'BU Code']

    #headers as in db
    headers_table = ['PetitionerName', 'PetitionerOfPrimaryBeneficiary', 'BeneficiaryXref', 'BeneficiaryType', 'Beneficiary_Xref2','FullName', 'SourceCreatedDate', 'CaseXref', 'CasePetitionName', 'CaseType', 'targetfiledate', 'CaseComments', 'LastStepCompleted', 'LastStepCompletedDate', 'SpecialInstructionFlag', 'SpecialInstructionInfo', 'DaysSinceLastStepCompleted', 'petitioningjobtitle', 'petitioningjoblocation', 'OnlineIntakeDate', 'questionnairesenttomanager', 'questionnairessenttofn', 'questionnairecompletedandreturnedbymanager', 'questionnairecompletedandreturnedbyfn', 'permmemosenttoemployer', 'approvalofpermmemoreceived', 'employeeworkexperiencechartsent', 'employeeworkexperiencechartreceived', 'employmentverificationletterssenttoemployee', 'signedemploymentverificationlettersreceived', 'prevailingwagedeterminationrequestsubmittedtodol', 'prevailingwagedeterminationissuedbydol', 'PwdExpirationDate', 'recruitmentinstructionssenttocompany', '_1stadditionalrecruitmentstepplaced', 'datedcopiesofallrecruitmentreceived', 'completedevaluationquestionnairesandresumesreceived', 'recruitmentreportsenttocompany', 'recruitmentreportreceived', 'form9089senttofnandemployer', 'editstoform9089receivedfromfnandemployer', 'form9089submittedtodol', 'Paralegal', 'SupervisoryParalegal','Attorney', 'BirthCountry', 'CitizenshipCountry', 'Current_Immigration_Status','CurrentImmigrationStatusExpirationDate2', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'FinalNivDate', 'PriorityDate1Date', 'EmployeeId', 'ManagerName', 'BusinessPartnerName', 'Department', 'Department_Group', 'Department_Number', 'CostCenter', 'CostCenterNumber', 'BusinessUnitCode']
        

    date_columns = ['SourceCreatedDate', 'targetfiledate', 'LastStepCompletedDate', 'OnlineIntakeDate', 'questionnairesenttomanager', 'questionnairessenttofn', 'questionnairecompletedandreturnedbymanager', 'questionnairecompletedandreturnedbyfn', 'permmemosenttoemployer', 'approvalofpermmemoreceived', 'employeeworkexperiencechartsent', 'employeeworkexperiencechartreceived', 'employmentverificationletterssenttoemployee', 'signedemploymentverificationlettersreceived', 'prevailingwagedeterminationrequestsubmittedtodol', 'PwdExpirationDate', 'prevailingwagedeterminationissuedbydol', 'recruitmentinstructionssenttocompany', 'datedcopiesofallrecruitmentreceived', 'completedevaluationquestionnairesandresumesreceived', 'recruitmentreportsenttocompany', 'recruitmentreportreceived', 'form9089senttofnandemployer', 'editstoform9089receivedfromfnandemployer', 'form9089submittedtodol', 'CurrentImmigrationStatusExpirationDate2', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'FinalNivDate', 'PriorityDate1Date']

    # print(date_columns)
    # quit()
    # header_names = [{'header': x} for x in headers]
    

    # # tab_1 CharterPermReport
    results_qry ='''select distinct  CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, p.PetitionerName,b.PetitionerOfPrimaryBeneficiary,b.BeneficiaryXref,b.BeneficiaryType,b.Beneficiary_Xref2,b.FullName,c.SourceCreatedDate,c.CaseXref,c.CasePetitionName,c.CaseType,c.targetfiledate,c.CaseComments,c.LastStepCompleted,c.LastStepCompletedDate,c.SpecialInstructionFlag,c.SpecialInstructionInfo,c.DaysSinceLastStepCompleted,c.petitioningjobtitle,c.petitioningjoblocation,c.OnlineIntakeDate,c.questionnairesenttomanager,c.questionnairessenttofn,c.questionnairecompletedandreturnedbymanager,c.questionnairecompletedandreturnedbyfn,c.permmemosenttoemployer,c.approvalofpermmemoreceived,c.employeeworkexperiencechartsent,c.employeeworkexperiencechartreceived,c.employmentverificationletterssenttoemployee,c.signedemploymentverificationlettersreceived,c.prevailingwagedeterminationrequestsubmittedtodol,c.prevailingwagedeterminationissuedbydol,c.PwdExpirationDate,c.recruitmentinstructionssenttocompany,c._1stadditionalrecruitmentstepplaced,c.datedcopiesofallrecruitmentreceived,c.completedevaluationquestionnairesandresumesreceived,c.recruitmentreportsenttocompany,c.recruitmentreportreceived,c.form9089senttofnandemployer,c.editstoform9089receivedfromfnandemployer,c.form9089submittedtodol,concat(c.ParalegalLastName,' ',c.ParalegalFirstName) as Paralegal,concat(c.SupervisoryParalegalLastName,' ', c.SupervisoryParalegalFirstName) as SupervisoryParalegal,concat(c.AssociateLastName,' ',c.AssociateFirstName) as Attorney,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.ImmigrationStatusExpirationDate,b.I797ExpirationDate,b.FinalNivDate,b.PriorityDate1Date,be.EmployeeId,be.ManagerName,be.BusinessPartnerName,be.Department,be.Department_Group,be.Department_Number,be.CostCenter,be.CostCenterNumber,be.BusinessUnitCode,b2.FullName as PrimBenFullName

    from [dbo].[Case] as c
    left join [dbo].[Beneficiary] as b on c.BeneficiaryXref = b.BeneficiaryXref 
    left join [dbo].[Petitioner] as p on p.PetitionerXref = b.PetitionerXref
    left join [dbo].[BeneficiaryEmployment] as be on be.BeneficiaryXref = b.BeneficiaryXref
    left join [dbo].[Beneficiary] as b2 on b.PrimaryBeneficiaryXref  = b2.BeneficiaryXref

    where b.IsActive=1 and 
    c.PrimaryCaseStatus = 'Open'  and
    c.CasePetitionId = '100003034' and

    lower(c.CaseDescription) in ('extension') and

    b.PetitionerOfPrimaryBeneficiary = 'Charter Communications, Inc.' and
    b.OrganizationXref = '100000590'
    order by PrimBenFullName asc ,  b.BeneficiaryType desc'''


    results = cursor.execute(results_qry).fetchall()

    df_tab1 = pd.read_sql(results_qry,conn)

    for dfcol in df_tab1.columns:
        if dfcol not in headers_table:
            df_tab1.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df_tab1 = df_tab1[headers_table]
    
    for d_h in date_columns:
        if d_h in df_tab1:
            if '1900-01-01' in df_tab1[d_h]:
                df_tab1[d_h] = ''
            else:
                df_tab1[d_h] = pd.to_datetime(df_tab1[d_h], format='%Y-%m-%d', errors='coerce').dt.date
   
    df_tab1.columns = headers #changing dataframe all column names


    # # ###########################################

    df_tab1.to_excel(writer, 'CharterPermReport', startrow=0, index=False)

    writer.save()
    writer.close()
    
    add_designs(file_path=result_filepath,date_months='')
  

def add_designs(file_path,date_months=''):
    
    exp_month = ((datetime.today())) + relativedelta(days=+270) 
    this_month = ((datetime.today()))
    next_month = ((datetime.today())) + relativedelta(months=+1) 

    end_month_str = pd.to_datetime(exp_month)
    this_month_str = pd.to_datetime(this_month)
    next_month_str = pd.to_datetime(next_month).date()

    # print ('\nAdding designs to the processed file..')
    book = load_workbook(file_path)
    for page_no,sheet in enumerate(book.sheetnames):
        page_no += 1
        # ws = book[book.sheetnames[sheet]] #for reading using sheet number
        ws = book.get_sheet_by_name(sheet)
        if ws:
            rows = ws.max_row 
            cols= ws.max_column 

            if rows<2:
                rows = 2
                ws.cell(row=2, column = 1).value = 'No Records Found'

            if sheet == 0:   
                ws.freeze_panes = ws['D2']
            else:
                ws.freeze_panes = ws['F2']

            for y in range(rows):
                for z in range(cols):

                    ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri (Body)', size = 11)

                    ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal='left', vertical='bottom')

                    ws.cell(row=y+1, column=z+1).font= Font(name = 'Calibri (Body)', size= 11)

                    if sheet.lower().__contains__('expiration'):  #
                        if str(ws.cell(row=1, column=z+1).value) in date_months:
                            doc_date = ws.cell(row=y+1, column=z+1).value

                            #old working conditions might be used later:
                            # try:
                            #     if (doc_date>=this_month) and (doc_date<=exp_month):
                                    
                            #         if(doc_date.month == this_month.month) and (doc_date.year == this_month.year):
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid') #red    
                            #         elif(doc_date.month == next_month.month) and (doc_date.year == next_month.year):
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type = 'solid') #orange
                            #         else:
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type = 'solid') #yellow
                            # except:
                            #     pass

                        # coloring blank cell condition:
                        #     if str(doc_date) in ['','nan','NaT','NaN','Nan','None']:
                        #         ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid') #red
                            
                            # #new conditions:
                            try:
                                if (doc_date<=exp_month):
                                    
                                    if(doc_date <= this_month):
                                        ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='F06969', end_color='F06969', fill_type = 'solid') #red    
                                    else:
                                        ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type = 'solid') #yellow
                            except:
                                pass

                    ws.cell(row=y+1, column=z+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    if y == 0:
                        ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri',size = 12, color = 'FFFFFF')

                        ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal='center', vertical='center')

        for cl in range(cols):
            if cl <= cols:
                ws.column_dimensions[get_column_letter(cl+1)].width = 15
                if (sheet == 'Acronyms'):
                    ws.column_dimensions[get_column_letter(cl+1)].width = 50
                    
        for rw in range(rows+1):
            if rw <= rows:
                ws.row_dimensions[rw].height = 30

        table = Table(displayName=f'Table_{page_no}', ref='A1:' + get_column_letter(cols) + str(rows))
        # .replace used above coz table display name cant hav spaces
        #     ws.cell(row=2, column = 1).value = 'No Records Found'

        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        table.tableStyleInfo = style
        ws.add_table(table)

        book.save(file_path)
    book.close()
 

# def add_designs(no_shts,file_path,doc_type='',date_months=''):
    
#     exp_month = ((datetime.today())) + relativedelta(days=+240) 
#     this_month = ((datetime.today()))
#     next_month = ((datetime.today())) + relativedelta(months=+1) 

#     end_month_str = pd.to_datetime(exp_month)
#     this_month_str = pd.to_datetime(this_month)
#     next_month_str = pd.to_datetime(next_month).date()



#     # print ('\nAdding designs to the processed file..')
#     book = load_workbook(file_path)
#     date_format = '%m/%d/%Y'
#     for sheet in range(no_shts):
#         ws = book[book.sheetnames[sheet]]
#         if ws:
#             rows = ws.max_row 
#             cols= ws.max_column 

#             if rows<2:
#                 rows = 2
#                 ws.cell(row=2, column = 1).value = 'No Records Found'

#             if sheet == 0:   
#                 ws.freeze_panes = ws['D2']
#             else:
#                 ws.freeze_panes = ws['F2']

#             for y in range(rows):
#                 for z in range(cols):

#                     ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri (Body)', size = 11)

#                     ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal='left', vertical='bottom')

#                     ws.cell(row=y+1, column=z+1).font= Font(name = 'Calibri (Body)', size= 11)

#                     if (sheet== 4 or sheet== 5) and  (doc_type == 'comcast'):  #
#                         if str(ws.cell(row=1, column=z+1).value) in date_months:
#                             doc_date = ws.cell(row=y+1, column=z+1).value

#                         #old working conditions might be used later:
#                             # try:
#                             #     if (doc_date>=this_month) and (doc_date<=exp_month):
                                    
#                             #         if(doc_date.month == this_month.month) and (doc_date.year == this_month.year):
#                             #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid') #red    
#                             #         elif(doc_date.month == next_month.month) and (doc_date.year == next_month.year):
#                             #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type = 'solid') #orange
#                             #         else:
#                             #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type = 'solid') #yellow
#                             # except:
#                             #     pass

#                         # coloring blank cell condition:

#                         #     if str(doc_date) in ['','nan','NaT','NaN','Nan','None']:
#                         #         ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid') #red
                            
#                             # #new conditions:
#                             try:
#                                 if (doc_date<=exp_month):
                                    
#                                     if(doc_date <= this_month):
#                                         ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='F06969', end_color='F06969', fill_type = 'solid') #red    
#                                     else:
#                                         ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type = 'solid') #yellow
#                             except:
#                                 pass

#                     ws.cell(row=y+1, column=z+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

#                     if y == 0:
#                         ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri',size = 12, color = 'FFFFFF')

#                         ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal='center', vertical='center')

#         for cl in range(cols):
#             if cl <= cols:
#                 ws.column_dimensions[get_column_letter(cl+1)].width = 15

#         for rw in range(rows+1):
#             if rw <= rows:
#                 ws.row_dimensions[rw].height = 30

#         table = Table(displayName='Table{}'.format(sheet+1), ref='A1:' + get_column_letter(cols) + str(rows))

#         #     ws.cell(row=2, column = 1).value = 'No Records Found'

#         style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)

#         table.tableStyleInfo = style
#         ws.add_table(table)

#         book.save(file_path)
#     book.close()
        
def PermReport(result_filepath2):
    # Tab 1 - PERM Report
    
    headers = ['Beneficiary Full Name','Management Info Department','Management Info Business Unit Code','Management Info Dept Group','Petitioning Job Title','Petitioning Job Location','Process ID','Date Opened','Birth Country','Citizenship','Current Status','NIV Max Out Date','Visa Priority Date','Online Intake Date','Questionnaire sent to manager','Questionnaires Sent to FN','Questionnaire completed and returned by Manager','Questionnaire completed and returned by FN','PERM Memo sent to employer','Approval of PERM Memo received','Employee Work Experience Chart sent','Employee Work Experience Chart received','Prevailing Wage Determination request submitted to DOL','Employment Verification Letters sent to employee','Prevailing Wage Determination issued by DOL','PWD Expiration Date','Recruitment instructions sent to company','1st Additional Recruitment Step Placed','Dated copies of all recruitment received','Completed evaluation questionnaires and resumes received','Recruitment report sent to company','Recruitment Report Received','Form 9089 sent to FN and Employer','Edits to Form 9089 received from FN and Employer','Form 9089 submitted to DOL','Date Closed','Days Since Last Activity']

    headers_table = ['FullName','Department','BusinessUnitCode','Department_Group','petitioningjobtitle','petitioningjoblocation','ParalegalXref','SourceCreatedDate','BirthCountry','CitizenshipCountry','Current_Immigration_Status','FinalNivDate','Priority1Date','OnlineIntakeDate','questionnairesenttomanager','questionnairessenttofn','questionnairecompletedandreturnedbymanager','questionnairecompletedandreturnedbyfn','permmemosenttoemployer','approvalofpermmemoreceived','employeeworkexperiencechartsent','employeeworkexperiencechartreceived','prevailingwagedeterminationrequestsubmittedtodol','employmentverificationletterssenttoemployee','prevailingwagedeterminationissuedbydol','PWDExpirationDate','recruitmentinstructionssenttocompany','_1stadditionalrecruitmentstepplaced','datedcopiesofallrecruitmentreceived','completedevaluationquestionnairesandresumesreceived','recruitmentreportsenttocompany','recruitmentreportreceived','form9089senttofnandemployer','editstoform9089receivedfromfnandemployer','form9089submittedtodol','CaseClosedDate','DaysSinceLastStepCompleted']
    
    date_columns = ['FinalNivDate','Priority1Date','SourceCreatedDate','OnlineIntakeDate','questionnairesenttomanager','questionnairessenttofn','questionnairecompletedandreturnedbymanager','questionnairecompletedandreturnedbyfn','permmemosenttoemployer','approvalofpermmemoreceived','employeeworkexperiencechartsent','employeeworkexperiencechartreceived','prevailingwagedeterminationrequestsubmittedtodol','employmentverificationletterssenttoemployee','prevailingwagedeterminationissuedbydol','PWDExpirationDate','recruitmentinstructionssenttocompany','_1stadditionalrecruitmentstepplaced','datedcopiesofallrecruitmentreceived','completedevaluationquestionnairesandresumesreceived','recruitmentreportsenttocompany','recruitmentreportreceived','form9089senttofnandemployer','editstoform9089receivedfromfnandemployer','form9089submittedtodol','CaseClosedDate']

    header_names = [{'header': x} for x in headers]
    
    results_active_qry = '''
        SELECT b.FullName,be.Department,be.BusinessUnitCode,be.Department_Group,c.petitioningjobtitle,c.petitioningjoblocation,c.ParalegalXref,b.BirthCountry,b.CitizenshipCountry,b.Current_Immigration_Status,b.FinalNivDate,bp.Priority1Date,c.SourceCreatedDate,c.OnlineIntakeDate,c.questionnairesenttomanager,c.questionnairessenttofn,c.questionnairecompletedandreturnedbymanager,c.questionnairecompletedandreturnedbyfn,c.permmemosenttoemployer,c.approvalofpermmemoreceived,c.employeeworkexperiencechartsent,c.employeeworkexperiencechartreceived,c.prevailingwagedeterminationrequestsubmittedtodol,c.employmentverificationletterssenttoemployee,c.prevailingwagedeterminationissuedbydol,c.PWDExpirationDate,c.recruitmentinstructionssenttocompany,c._1stadditionalrecruitmentstepplaced,c.datedcopiesofallrecruitmentreceived,c.completedevaluationquestionnairesandresumesreceived,c.recruitmentreportsenttocompany,c.recruitmentreportreceived,c.form9089senttofnandemployer,c.editstoform9089receivedfromfnandemployer,c.form9089submittedtodol,c.CaseClosedDate,c.DaysSinceLastStepCompleted,o.OrganizationXref,
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus
        FROM dbo.[Case] as c  
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.BeneficiaryEmployment as be on be.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.Organization as o on b.OrganizationXref = o.OrganizationXref
        where b.IsActive = '1' and c.PrimaryCaseStatus='Open' and o.OrganizationXref = '100000590' and
        (c.CaseType = 'Labor Cert PERM' or c.CaseType = 'Labor Cert Special Handling' )
        ORDER BY b.FullName ASC'''
        # filter this using org id instead pertitioner id
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if '1900-01-01' in df[d_h]:
                df[d_h] = ''
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names

    writer = pd.ExcelWriter(result_filepath2, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, 'PERM Report', startrow=0, columns=headers, index=False)
    writer.save()
    
    add_designs(file_path=result_filepath2,date_months='')


def  WeeklyNetOps(result_filepath3):
    ###################################### Tab 1 Header #############################################
    # Tab 1 - Weekly NetOps - P&T Transfer Report
    

    headers = ['Beneficiary Full Name','Management Info Employee ID','Management Info Department','Management Info Business Unit Code','Management Info Dept Group','Management Info Job Start Date','Birth Country','Citizenship','Petitioning Job Title','Petitioning Job Location','Current Status','Current Status Expiration Date','I-797 Expiration Date','I-94 Expiration Date','EAD Expiration','AP Expiration','Management Info Manager','Management Info Second Level Manager','NIV Max Out Date','Visa Priority Date','Special Instruction Flag','Case Opened','Process Type','Process Reference','Questionnaires Sent to FN','FN Completed Questionnaires and Acknowledgement','All FN Docs Received','LCA Filed','LCA Certified','Forms and Documentation Submitted for Signature','Petition Filed with CIS','Receipt Date','RFE Received','RFE Due Date','RFE Docs Received','RFE Response Submitted','Final Action','Date Status Valid From','Date Status Valid To','Case Closed','Days Since Last Activity']

    headers_table = ['FullName','EmployeeId','Department','BusinessUnitCode','Department_Group','EmploymentStartDate','BirthCountry','CitizenshipCountry','petitioningjobtitle','petitioningjoblocation','Current_Immigration_Status','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','EadExpirationDate','AdvanceParoleExpirationDate','ManagerName','SecondLevelManager','FinalNivDate','Priority1Date','SpecialInstructionFlag','SourceCreatedDate','CaseType','CaseDescription','questionnairessenttofn','fncompletedquestionnairesandacknowledgement','allfndocsreceived','lcafiled','lcacertified','formsanddocumentationsubmittedforsignature','petitionfiledwithcis','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEDocsReceivedDate','RFEAuditSubmittedDate','SecondaryCaseStatus','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','DaysSinceLastStepCompleted']

    date_columns = ['EmploymentStartDate','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','EadExpirationDate','AdvanceParoleExpirationDate','FinalNivDate','Priority1Date','SourceCreatedDate','questionnairessenttofn','fncompletedquestionnairesandacknowledgement','allfndocsreceived','lcafiled','lcacertified','formsanddocumentationsubmittedforsignature','applicationfiled','petitionfiledwithcis','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEDocsReceivedDate','RFEAuditSubmittedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate']


    header_names = [{'header': x} for x in headers]
    
    results_active_qry = ''' SELECT 
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 
        b.FullName,be.EmployeeId,be.Department,be.BusinessUnitCode,be.Department_Group,be.EmploymentStartDate,b.BirthCountry,b.CitizenshipCountry,c.petitioningjobtitle,c.petitioningjoblocation,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.I797ExpirationDate,b.ImmigrationStatusExpirationDate,b.EadExpirationDate,b.AdvanceParoleExpirationDate,be.ManagerName,be.SecondLevelManager,b.FinalNivDate,bp.Priority1Date,c.SpecialInstructionFlag,c.SourceCreatedDate,c.CaseDescription,c.CaseType,c.questionnairessenttofn,c.fncompletedquestionnairesandacknowledgement,c.allfndocsreceived,c.lcafiled,c.lcacertified,c.formsanddocumentationsubmittedforsignature,c.applicationfiled,c.petitionfiledwithcis,c.CaseReceivedDate,c.RFEAuditReceivedDate,c.RFEAuditDueDate,c.RFEDocsReceivedDate,c.RFEAuditSubmittedDate,c.SecondaryCaseStatus,c.CaseValidFromDate,c.CaseExpirationDate,c.CaseClosedDate,c.DaysSinceLastStepCompleted
        FROM dbo.[Case] as c  
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.BeneficiaryEmployment as be on be.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryXref=b.BeneficiaryXref
        LEFT JOIN dbo.[Organization] as o on b.OrganizationXref =o.OrganizationXref
        where b.IsActive = '1' and c.PrimaryCaseStatus='Open' and (c.CaseType='Change of Employer' or c.CaseType='New Hire Assessment')
        and o.OrganizationXref = '100000590'
        ORDER BY b.FullName ASC '''
        # orgnaniastion 
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if '1900-01-01' in df[d_h]:
                df[d_h] = ''
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath3, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, 'Weekly NetOps-P&T', startrow=0, columns=headers, index=False)
    writer.save()
    
    add_designs(file_path=result_filepath3,date_months='')


def Paralegal(result_filepath4):
    
    headers = ['Beneficiary Name', 'Case No', 'Petitioner', 'Birth Country', 'Current Status', 'Current Status Expiration Date', 'I-94 Expiration Date', 'I-797 Expiration Date', 'NIV Max Out Date', 'EAD Expiration', 'AP Expiration',  'HR Special Instructions Flag','Special Instruction Flag',  'Process Id','Date Opened', 'Process Type', 'Process Reference', 'Target File Date', 'Summary Case Disposition', 'Last Process Activity', 'Last Process Activity Date', 'Days Since Last Activity', 'Next Unfinished Reminder - Subject', 'Next Unfinished Reminder - Expiry', 'Visa Priority Date', 'Visa Preference', 'Visa Priority Note','Case Filed', 'Petition Filed with CIS', 'Form I-129 Filed with CIS', 'Form 9089 submitted to DOL', 'AOS Application Filed', 'RFE Due Date', 'RFE Response Submitted', 'Associate Name', 'Paralegal Name'] 

    headers_table = ['FullName', 'Beneficiary_Xref2', 'PetitionerofPrimaryBeneficiary', 'BirthCountry', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'FinalNivDate', 'EadExpirationDate', 'AdvanceParoleExpirationDate', 'SpecialInstructionInfo', 'SpecialInstructionFlag',  'CaseXref', 'SourceCreatedDate','CaseType', 'CaseDescription','targetfiledate', 'CaseComments', 'LastStepCompleted', 'LastStepCompletedDate', 'DaysSinceLastStepCompleted', 'NextStepAction', 'NextStepActionDueDate', 'PriorityDate1Date', 'PriorityDate1Category', 'PriorityDate1Note', 'CaseFiledDate','petitionfiledwithcis',  'formi129filedwithcis', 'form9089submittedtodol', 'aosapplicationfiled', 'RFEAuditDueDate', 'RFEAuditSubmittedDate', 'AssociateName','ParalegalName']


    date_columns = ['EmploymentStartDate','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','EadExpirationDate','AdvanceParoleExpirationDate','FinalNivDate','Priority1Date','SourceCreatedDate','questionnairessenttofn','fncompletedquestionnairesandacknowledgement','allfndocsreceived','lcafiled','lcacertified','formsanddocumentationsubmittedforsignature','applicationfiled','petitionfiledwithcis','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEDocsReceivedDate','RFEAuditSubmittedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','LastStepCompletedDate','NextStepActionDueDate','applicationfiled','CaseFiledDate','targetfiledate','PriorityDate1Date','formi129filedwithcis','form9089submittedtodol','aosapplicationfiled']


    header_names = [{'header': x} for x in headers]
    
    results_active_qry = ''' SELECT 
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 
        b.FullName, b.Beneficiary_Xref2, b.PetitionerofPrimaryBeneficiary, b.BirthCountry, b.Current_Immigration_Status, b.CurrentImmigrationStatusExpirationDate2, b.ImmigrationStatusExpirationDate, b.I797ExpirationDate, b.FinalNivDate, b.EadExpirationDate, b.AdvanceParoleExpirationDate, p.SpecialInstructionInfo, p.SpecialInstructionFlag, p.SourceCreatedDate, p.CaseXref, p.CaseDescription, p.CaseType, p.targetfiledate, p.CaseComments, p.LastStepCompleted, p.LastStepCompletedDate, p.DaysSinceLastStepCompleted, p.NextStepAction, 
        p.NextStepActionDueDate, b.PriorityDate1Date, b.PriorityDate1Category, b.PriorityDate1Note, p.petitionfiledwithcis, p.CaseFiledDate, p.applicationfiled, p.formi129filedwithcis, p.form9089submittedtodol, p.aosapplicationfiled, p.RFEAuditDueDate, p.RFEAuditSubmittedDate, concat(p.AssociateLastName,', ',p.AssociateFirstName) as AssociateName, concat(p.ParalegalLastName,', ',p.ParalegalFirstName) as ParalegalName

        FROM dbo.[Case] as p  
        LEFT JOIN  dbo.Beneficiary as b on p.BeneficiaryXref = b.BeneficiaryXref
        where b.IsActive = '1' and p.PrimaryCaseStatus='Open'
        ORDER BY b.FullName ASC '''
        
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if '1900-01-01' in df[d_h]:
                df[d_h] = ''
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath4, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, 'Paralegal Active List', startrow=0, columns=headers, index=False)
    writer.save()
    
    add_designs(file_path=result_filepath4,date_months='')

def InternalPerm(result_filepath5):

    headers = ['Beneficiary Name', 'Case No', 'Petitioner', 'Birth Country', 'Current Status', 'Current Status Expiration Date', 'I-94 Expiration Date', 'I-797 Expiration Date', 'NIV Max Out Date', 'EAD Expiration', 'AP Expiration',  'HR Special Instructions Flag', 'Special Instruction Flag', 'Process Id', 'Process  Opened Date','Process Type', 'Process Reference', 'Target File Date', 'Summary Case Disposition', 'Last Process Activity','Primary Process Status', 'Last Process Activity Date', 'Days Since Last Activity', 'Next Unfinished Reminder - Subject', 'Next Unfinished Reminder - Expiry', 'Petitioning Job Title', 'Petitioning Job Location', 'Questionnaires Sent to FN', 'Follow up with FN for requested information', 'Questionnaire completed and returned by Manager', 'Questionnaire completed and returned by FN', 'PERM Memo sent to employer', 'Approval of PERM Memo received', 'Employee Work Experience Chart sent', 'Employee Work Experience Chart received', 'Employment Verification Letters sent to employee', 'Signed Employment Verification Letters received', 'Prevailing Wage Determination request submitted to DOL', 'Prevailing Wage Determination issued by DOL', 'Recruitment instructions sent to company', 'Job Order Placed with SWA', 'Notice of Filing posted', 'Intranet Notice of Filing Posted', 'Notice of Filing removed signed', 'Intranet Notice of Filing Removed', '1st Sunday Ad Placed', '2nd Sunday Ad Placed', '1st Additional Recruitment Step Placed', '2nd Additional Recruitment Step Placed', '3rd Additional Recruitment Step Placed', 'Dated copies of all recruitment received', 'Completed evaluation questionnaires and resumes received', 'Form 9089 sent to FN and Employer', 'Edits to Form 9089 received from FN and Employer', 'Form 9089 submitted to DOL', 'Associate Name', 'Paralegal Name']

    headers_table = ['FullName', 'Beneficiary_Xref2', 'PetitionerofPrimaryBeneficiary', 'BirthCountry', 'Current_Immigration_Status', 'CurrentImmigrationStatusExpirationDate2', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'FinalNivDate', 'EadExpirationDate', 'AdvanceParoleExpirationDate', 'SpecialInstructionInfo', 'SpecialInstructionFlag',  'CaseXref','SourceCreatedDate','CaseType', 'CaseDescription','targetfiledate', 'CaseComments', 'LastStepCompleted','PrimaryCaseStatus', 'LastStepCompletedDate', 'DaysSinceLastStepCompleted', 'NextStepAction', 'NextStepActionDueDate', 'petitioningjobtitle', 'petitioningjoblocation', 'questionnairessenttofn', 'followupwithfnforrequestedinformation', 'questionnairecompletedandreturnedbymanager', 'questionnairecompletedandreturnedbyfn', 'permmemosenttoemployer', 'approvalofpermmemoreceived', 'employeeworkexperiencechartsent', 'employeeworkexperiencechartreceived', 'employmentverificationletterssenttoemployee', 'signedemploymentverificationlettersreceived', 'prevailingwagedeterminationrequestsubmittedtodol', 'prevailingwagedeterminationissuedbydol', 'recruitmentinstructionssenttocompany', 'joborderplacedwithswa', 'noticeoffilingposted', 'intranetnoticeoffilingposted', 'noticeoffilingremovedsigned', 'intranetnoticeoffilingremoved', '_1stsundayadplaced', '_2ndsundayadplaced', '_1stadditionalrecruitmentstepplaced', '_2ndadditionalrecruitmentstepplaced', '_3rdadditionalrecruitmentstepplaced', 'datedcopiesofallrecruitmentreceived', 'completedevaluationquestionnairesandresumesreceived', 'form9089senttofnandemployer', 'editstoform9089receivedfromfnandemployer', 'form9089submittedtodol', 'AssociateName', 'ParalegalName']
    
    date_columns = ['targetfiledate','noticeoffilingremovedsigned', 'intranetnoticeoffilingremoved','intranetnoticeoffilingposted','noticeoffilingposted','signedemploymentverificationlettersreceived','joborderplacedwithswa','EmploymentStartDate','CurrentImmigrationStatusExpirationDate2','I797ExpirationDate','ImmigrationStatusExpirationDate','EadExpirationDate','AdvanceParoleExpirationDate','FinalNivDate','Priority1Date','SourceCreatedDate','questionnairessenttofn','fncompletedquestionnairesandacknowledgement','allfndocsreceived','lcafiled','lcacertified','formsanddocumentationsubmittedforsignature','applicationfiled','petitionfiledwithcis','CaseReceivedDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEDocsReceivedDate','RFEAuditSubmittedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','LastStepCompletedDate','NextStepActionDueDate','applicationfiled','CaseFiledDate','form9089submittedtodol','editstoform9089receivedfromfnandemployer','form9089senttofnandemployer','_1stsundayadplaced', '_2ndsundayadplaced','_1stadditionalrecruitmentstepplaced','_2ndadditionalrecruitmentstepplaced','_3rdadditionalrecruitmentstepplaced', 'datedcopiesofallrecruitmentreceived','completedevaluationquestionnairesandresumesreceived','FinalNivDate','Priority1Date','SourceCreatedDate','OnlineIntakeDate','questionnairesenttomanager','questionnairessenttofn','questionnairecompletedandreturnedbymanager','questionnairecompletedandreturnedbyfn','permmemosenttoemployer','approvalofpermmemoreceived','employeeworkexperiencechartsent','employeeworkexperiencechartreceived','prevailingwagedeterminationrequestsubmittedtodol','employmentverificationletterssenttoemployee','prevailingwagedeterminationissuedbydol','PWDExpirationDate','recruitmentinstructionssenttocompany','_1stadditionalrecruitmentstepplaced','datedcopiesofallrecruitmentreceived','completedevaluationquestionnairesandresumesreceived','recruitmentreportsenttocompany','recruitmentreportreceived','form9089senttofnandemployer','editstoform9089receivedfromfnandemployer','form9089submittedtodol','CaseClosedDate']


    header_names = [{'header': x} for x in headers]
    
    results_active_qry = ''' SELECT 
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Retired' END as BeneficiaryRecordStatus, 

        b.FullName,b.Beneficiary_Xref2,b.PetitionerofPrimaryBeneficiary,b.BirthCountry,b.Current_Immigration_Status,b.CurrentImmigrationStatusExpirationDate2,b.ImmigrationStatusExpirationDate,b.I797ExpirationDate,b.FinalNivDate,b.EadExpirationDate,b.AdvanceParoleExpirationDate,p.SpecialInstructionInfo,p.SpecialInstructionFlag,p.SourceCreatedDate,p.CaseXref,p.CasePetitionName,p.CaseType,p.CaseDescription,p.targetfiledate,p.CaseComments,p.LastStepCompleted,p.LastStepCompletedDate,p.DaysSinceLastStepCompleted,p.NextStepAction,p.NextStepActionDueDate,p.petitioningjobtitle,p.petitioningjoblocation,p.questionnairessenttofn,p.followupwithfnforrequestedinformation,p.questionnairecompletedandreturnedbymanager,p.PrimaryCaseStatus,p.questionnairecompletedandreturnedbyfn,p.permmemosenttoemployer,p.approvalofpermmemoreceived,p.employeeworkexperiencechartsent,p.employeeworkexperiencechartreceived,p.employmentverificationletterssenttoemployee,p.signedemploymentverificationlettersreceived,p.prevailingwagedeterminationrequestsubmittedtodol,p.prevailingwagedeterminationissuedbydol,p.recruitmentinstructionssenttocompany,p.joborderplacedwithswa,p.noticeoffilingposted,p.intranetnoticeoffilingposted,p.noticeoffilingremovedsigned,p.intranetnoticeoffilingremoved,p._1stsundayadplaced,p._2ndsundayadplaced,p._1stadditionalrecruitmentstepplaced,p._2ndadditionalrecruitmentstepplaced,p._3rdadditionalrecruitmentstepplaced,p.datedcopiesofallrecruitmentreceived,p.completedevaluationquestionnairesandresumesreceived,p.form9089senttofnandemployer,p.editstoform9089receivedfromfnandemployer,p.form9089submittedtodol,concat(p.AssociateLastName,', ',p.AssociateFirstName) as AssociateName,concat(p.ParalegalLastName,', ',p.ParalegalFirstName) as ParalegalName

        FROM dbo.[Case] as p  
        LEFT JOIN  dbo.Beneficiary as b on p.BeneficiaryXref = b.BeneficiaryXref
        where b.IsActive = '1' and p.PrimaryCaseStatus='Open' and
        (p.CaseType = 'Labor Cert PERM' or p.CaseType = 'Labor Cert Special Handling' )
        ORDER BY b.FullName ASC	 '''
        
    results_active = cursor.execute(results_active_qry).fetchall()
    
    df = pd.read_sql(results_active_qry, conn)
    for dfcol in df.columns:
        if dfcol not in headers_table :
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]
    for d_h in date_columns:
        if d_h in df:
            if '1900-01-01' in df[d_h]:
                df[d_h] = ''
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    writer = pd.ExcelWriter(result_filepath5, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, 'Internal-PERM Report', startrow=0, columns=headers, index=False)
    writer.save()
    
    add_designs(file_path=result_filepath5,date_months='')


if __name__ == '__main__':
    print('Program Execution Started..\nIn Progress..\n')
    start()
    # truncate_full_db()
    print('Finished')
    pass
    
    
    

